# © 2025 HarmonyCares
# All rights reserved.
# ruff: noqa
# ruff: noqa

"""
Excel Multi-Sheet Analyzer - Development Tool

A comprehensive analysis tool for understanding the structure of multi-sheet Excel files.
Uses pylightxl for fast, lightweight Excel reading without pandas dependency.

This tool helps developers create accurate schemas by revealing:
- Complete cell-by-cell data and formulas
- Data patterns and section boundaries
- Potential header rows and data regions
- Non-empty cell distributions
- Field extraction opportunities

Or via CLI:
    python -m acoharmony._dev.excel_analyzer path/to/file.xlsx [--output analysis.json]
"""

import json
import sys
from pathlib import Path
from collections import defaultdict
from typing import Any, Dict, List, Optional
from dataclasses import dataclass, asdict

try:
    import pylightxl as xl  # pragma: no cover
    HAS_PYLIGHTXL = True  # pragma: no cover
except ImportError:
    HAS_PYLIGHTXL = False
    xl = None


@dataclass
class CellInfo:
    """Information about a single cell."""
    row: int
    col: int
    value: Any
    value_type: str  # 'string', 'number', 'boolean', 'null', 'error'
    matrix_notation: str  # [sheet, row, col]

    def to_dict(self) -> dict:
        return {
            "row": self.row,
            "col": self.col,
            "value": str(self.value) if self.value is not None else None,  # FULL value, no truncation
            "value_type": self.value_type,
            "matrix_notation": self.matrix_notation
        }


@dataclass
class DataSection:
    """A contiguous section of data within a sheet."""
    start_row: int
    end_row: int
    row_count: int
    avg_non_null_per_row: float

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class PotentialHeader:
    """A row that might be a header."""
    row: int
    non_null_count: int
    values: List[str]
    confidence: str  # 'high', 'medium', 'low'

    def to_dict(self) -> dict:
        return {
            "row": self.row,
            "non_null_count": self.non_null_count,
            "values": self.values,  # ALL values, no truncation
            "confidence": self.confidence
        }


@dataclass
class SheetAnalysis:
    """Complete analysis of a single sheet."""
    sheet_index: int
    sheet_name: str
    dimensions: Dict[str, int]
    non_empty_cell_count: int
    cells: List[CellInfo]
    potential_headers: List[PotentialHeader]
    data_sections: List[DataSection]
    column_stats: Dict[int, Dict[str, Any]]
    value_type_distribution: Dict[str, int]

    def to_dict(self) -> dict:
        return {
            "sheet_index": self.sheet_index,
            "sheet_name": self.sheet_name,
            "dimensions": self.dimensions,
            "non_empty_cell_count": self.non_empty_cell_count,
            "cells": [c.to_dict() for c in self.cells],
            "potential_headers": [h.to_dict() for h in self.potential_headers],
            "data_sections": [s.to_dict() for s in self.data_sections],
            "column_stats": self.column_stats,
            "value_type_distribution": self.value_type_distribution
        }


class ExcelAnalyzer:
    """
    Comprehensive Excel workbook analyzer using pylightxl.

        Analyzes all sheets in an Excel file and provides detailed structural information
        to help developers understand data layout and create accurate parsing schemas.

        Parameters

        file_path : str or Path
            Path to Excel file to analyze
        max_cell_value_length : int, default=200
            Maximum length for cell values in output (longer values are truncated)
        include_empty_cells : bool, default=False
            Whether to include empty cells in cell-level analysis
    """

    def __init__(
        self,
        file_path: str | Path,
        max_cell_value_length: int = 200,
        include_empty_cells: bool = False
    ):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        self.max_cell_value_length = max_cell_value_length
        self.include_empty_cells = include_empty_cells

        # Load workbook
        self.db = xl.readxl(fn=str(self.file_path))

        # Get sheet names in their actual workbook order (not alphabetical)
        # pylightxl.ws_names returns alphabetical order, so use openpyxl for correct order
        try:  # pragma: no cover – requires pylightxl
            from openpyxl import load_workbook
            wb_temp = load_workbook(str(self.file_path), read_only=True, data_only=True)
            self.sheet_names = wb_temp.sheetnames
            wb_temp.close()
        except ImportError:  # pragma: no cover
            # Fallback to pylightxl order if openpyxl not available
            self.sheet_names = self.db.ws_names

        self.analysis: Dict[str, Any] = {}

    def _classify_value_type(self, value: Any) -> str:
        """Classify the type of a cell value."""
        if value is None or value == '':
            return "null"

        value_str = str(value)

        # Check for Excel formula errors
        if value_str.startswith('#'):
            return "error"

        # Try to determine type
        if isinstance(value, bool):
            return "boolean"
        elif isinstance(value, (int, float)):
            return "number"
        else:
            return "string"

    def _analyze_potential_headers(self, ws, max_rows: int = 20) -> List[PotentialHeader]:
        """
        Identify rows that might be headers based on data patterns.

                Headers typically have:
                - Many non-null values
                - String values
                - Different pattern than subsequent rows
        """
        potential_headers = []
        size = ws.size

        for row_idx in range(1, min(max_rows + 1, size[0] + 1)):
            row_data = ws.row(row_idx)
            non_null_values = [v for v in row_data if v is not None and v != '']
            non_null_count = len(non_null_values)

            # Skip rows with very few values
            if non_null_count < 2:
                continue

            # Get string representations - FULL values
            non_null_str = [str(v) for v in non_null_values]

            # Determine confidence based on patterns
            confidence = "low"
            if non_null_count >= size[1] * 0.7:  # 70% or more columns filled
                confidence = "high"
            elif non_null_count >= size[1] * 0.4:  # 40-70% filled
                confidence = "medium"

            # Higher confidence if values look like header text
            if any(len(str(v)) > 5 and not str(v).replace('.', '').replace('-', '').isdigit()
                   for v in non_null_values[:5]):
                if confidence == "low":
                    confidence = "medium"
                elif confidence == "medium":
                    confidence = "high"

            potential_headers.append(PotentialHeader(
                row=row_idx - 1,  # Convert to 0-indexed
                non_null_count=non_null_count,
                values=non_null_str,  # ALL values
                confidence=confidence
            ))

        return potential_headers

    def _identify_data_sections(self, ws) -> List[DataSection]:
        """
        Identify contiguous sections of data (separated by empty rows).
        """
        sections = []
        current_section_start = None
        section_row_counts = []
        size = ws.size

        for row_idx in range(1, size[0] + 1):
            row_data = ws.row(row_idx)
            row_has_data = any(v is not None and v != '' for v in row_data)

            if row_has_data:
                if current_section_start is None:
                    current_section_start = row_idx
                    section_row_counts = []

                # Track non-null count for this row
                non_null_count = sum(1 for v in row_data if v is not None and v != '')
                section_row_counts.append(non_null_count)

            elif current_section_start is not None:
                # End of section
                avg_non_null = sum(section_row_counts) / len(section_row_counts) if section_row_counts else 0
                sections.append(DataSection(
                    start_row=current_section_start - 1,  # Convert to 0-indexed
                    end_row=row_idx - 2,  # Convert to 0-indexed
                    row_count=row_idx - current_section_start,
                    avg_non_null_per_row=round(avg_non_null, 2)
                ))
                current_section_start = None
                section_row_counts = []

        # Close last section if still open
        if current_section_start is not None:
            avg_non_null = sum(section_row_counts) / len(section_row_counts) if section_row_counts else 0
            sections.append(DataSection(
                start_row=current_section_start - 1,  # Convert to 0-indexed
                end_row=size[0] - 1,
                row_count=size[0] - current_section_start + 1,
                avg_non_null_per_row=round(avg_non_null, 2)
            ))

        return sections

    def _analyze_column_stats(self, ws) -> Dict[int, Dict[str, Any]]:
        """Analyze statistics for each column."""
        stats = {}
        size = ws.size

        for col_idx in range(1, size[1] + 1):
            col_data = ws.col(col_idx)

            non_null_values = [v for v in col_data if v is not None and v != '']
            non_null_count = len(non_null_values)

            if non_null_count == 0:
                continue

            # Get value types
            value_types = defaultdict(int)
            sample_values = []

            for val in non_null_values:  # ALL non-null values
                val_type = self._classify_value_type(val)
                value_types[val_type] += 1

            # Store ALL sample values, not just first 10
            sample_values = [str(val) for val in non_null_values]

            stats[col_idx - 1] = {  # Convert to 0-indexed
                "non_null_count": non_null_count,
                "null_count": len(col_data) - non_null_count,
                "fill_rate": round(non_null_count / len(col_data), 3),
                "value_types": dict(value_types),
                "sample_values": sample_values  # ALL values
            }

        return stats

    def analyze_sheet(self, sheet_index: int, sheet_name: str) -> SheetAnalysis:
        """
        Analyze a single sheet comprehensively.

                Parameters

                sheet_index : int
                    0-based sheet index
                sheet_name : str
                    Name of the sheet

                Returns

                SheetAnalysis
                    Complete analysis of the sheet
        """
        ws = self.db.ws(ws=sheet_name)
        size = ws.size  # (rows, cols)

        # Extract all cells
        cells = []
        value_type_counts = defaultdict(int)

        for row_idx in range(1, size[0] + 1):
            for col_idx in range(1, size[1] + 1):
                cell_value = ws.index(row=row_idx, col=col_idx)
                value_type = self._classify_value_type(cell_value)
                value_type_counts[value_type] += 1

                # Skip empty cells unless requested
                if not self.include_empty_cells and value_type == "null":
                    continue

                if cell_value is not None or self.include_empty_cells:  # pragma: no branch
                    cell_info = CellInfo(
                        row=row_idx - 1,  # Convert to 0-indexed
                        col=col_idx - 1,  # Convert to 0-indexed
                        value=cell_value,
                        value_type=value_type,
                        matrix_notation=f"[{sheet_index}, {row_idx - 1}, {col_idx - 1}]"
                    )
                    cells.append(cell_info)

        # Identify potential headers
        potential_headers = self._analyze_potential_headers(ws)

        # Identify data sections
        data_sections = self._identify_data_sections(ws)

        # Column statistics
        column_stats = self._analyze_column_stats(ws)

        return SheetAnalysis(
            sheet_index=sheet_index,
            sheet_name=sheet_name,
            dimensions={"rows": size[0], "cols": size[1]},
            non_empty_cell_count=len([c for c in cells if c.value_type != "null"]),
            cells=cells,
            potential_headers=potential_headers,
            data_sections=data_sections,
            column_stats=column_stats,
            value_type_distribution=dict(value_type_counts)
        )

    def analyze(self, sheets: Optional[List[int]] = None) -> Dict[str, Any]:
        """
        Analyze all sheets (or specified sheets) in the workbook.

                Parameters

                sheets : List[int], optional
                    List of sheet indices to analyze. If None, analyzes all sheets.

                Returns

                Dict[str, Any]
                    Complete analysis results
        """
        sheet_indices = sheets if sheets is not None else range(len(self.sheet_names))

        self.analysis = {
            "file_name": self.file_path.name,
            "file_path": str(self.file_path),
            "total_sheets": len(self.sheet_names),
            "sheet_names": self.sheet_names,
            "sheets": {},
            "summary": {}
        }

        total_cells = 0
        total_non_empty = 0

        for sheet_idx in sheet_indices:
            if sheet_idx >= len(self.sheet_names):
                print(f"Warning: Sheet index {sheet_idx} out of range, skipping")
                continue

            sheet_name = self.sheet_names[sheet_idx]

            try:
                sheet_analysis = self.analyze_sheet(sheet_idx, sheet_name)

                key = f"sheet_{sheet_idx}_{sheet_name}"
                self.analysis["sheets"][key] = sheet_analysis

                total_cells += sheet_analysis.dimensions["rows"] * sheet_analysis.dimensions["cols"]
                total_non_empty += sheet_analysis.non_empty_cell_count

            except Exception as e:  # ALLOWED: Dev tool - prints error, stores error in result, continues with remaining sheets
                print(f"Error analyzing sheet {sheet_idx} ({sheet_name}): {e}")
                self.analysis["sheets"][f"sheet_{sheet_idx}_{sheet_name}"] = {
                    "error": str(e)
                }

        self.analysis["summary"] = {
            "total_cells": total_cells,
            "total_non_empty_cells": total_non_empty,
            "fill_rate": round(total_non_empty / total_cells, 3) if total_cells > 0 else 0
        }

        return self.analysis

    def save_json(self, output_path: str | Path) -> None:
        """Save analysis results to JSON file."""
        if not self.analysis:
            raise ValueError("No analysis results to save. Run analyze() first.")

        # Convert to JSON-serializable format
        json_data = {
            "file_name": self.analysis["file_name"],
            "file_path": self.analysis["file_path"],
            "total_sheets": self.analysis["total_sheets"],
            "sheet_names": self.analysis["sheet_names"],
            "summary": self.analysis["summary"],
            "sheets": {}
        }

        for key, sheet_analysis in self.analysis["sheets"].items():
            if isinstance(sheet_analysis, SheetAnalysis):
                json_data["sheets"][key] = sheet_analysis.to_dict()
            else:
                json_data["sheets"][key] = sheet_analysis

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w') as f:
            json.dump(json_data, f, indent=2)

        print(f"Analysis saved to: {output_path}")

    def save_to_dev_logs(self, base_name: str = None) -> Path:
        """
        Save full workbook analysis to logs/dev directory using storage backend.

                Parameters

                base_name : str, optional
                    Base name for the output file. If None, uses the Excel filename.

                Returns

                Path
                    Path to the saved analysis file
        """
        if not self.analysis:
            raise ValueError("No analysis results to save. Run analyze() first.")

        # Get storage backend paths
        try:
            from .._store import StorageBackend
            import os

            profile = os.getenv("ACO_PROFILE", "local")
            storage = StorageBackend(profile=profile)
            logs_dir = storage.get_path("logs")
        except Exception as e:
            from .._exceptions import StorageBackendError
            import os
            profile = os.getenv("ACO_PROFILE")
            raise StorageBackendError.from_initialization_error(e, profile)

        # Create logs/dev directory
        dev_logs_dir = logs_dir / "dev"
        dev_logs_dir.mkdir(parents=True, exist_ok=True)

        # Generate output filename
        if base_name is None:
            base_name = self.file_path.stem

        output_path = dev_logs_dir / f"{base_name}_analysis.json"

        # Save using existing save_json method
        self.save_json(output_path)

        return output_path

    def print_summary(self, max_sheets: Optional[int] = None) -> None:
        """Print a direct, actionable summary focused on what matters for schema creation."""
        if not self.analysis:
            print("No analysis results. Run analyze() first.")
            return

        print("=" * 100)
        print(f"EXCEL ANALYSIS: {self.analysis['file_name']}")
        print("=" * 100)
        print(f"\nTotal Sheets: {self.analysis['total_sheets']}")
        print(f"Sheet Names: {', '.join(self.analysis['sheet_names'])}")

        print(f"\n{'=' * 100}")
        print("SHEET STRUCTURE (what you need to know)")
        print("=" * 100)

        sheet_count = 0
        for key, sheet_data in self.analysis["sheets"].items():
            if max_sheets and sheet_count >= max_sheets:
                print(f"\n... {len(self.analysis['sheets']) - max_sheets} more sheets ...")
                break

            sheet_count += 1

            if isinstance(sheet_data, SheetAnalysis):
                sheet_data = sheet_data.to_dict()

            if "error" in sheet_data:
                print(f"\n[FAILED] Sheet {key}: ERROR - {sheet_data['error']}")
                continue

            print(f"\n📄 Sheet {sheet_data['sheet_index']}: {sheet_data['sheet_name']}")
            print(f"   Size: {sheet_data['dimensions']['rows']} rows × {sheet_data['dimensions']['cols']} cols")

            # Show actual header row with values
            if sheet_data['potential_headers']:
                best_header = sheet_data['potential_headers'][0]
                print(f"\n   🎯 LIKELY HEADER at row {best_header['row']} ({best_header['confidence']} confidence):")
                print(f"      Values: {best_header['values']}")

                # Show other potential headers if they exist
                if len(sheet_data['potential_headers']) > 1:
                    print(f"\n   Other potential headers:")
                    for h in sheet_data['potential_headers'][1:4]:
                        print(f"      Row {h['row']:3d}: {h['values'][:5]}")

            # Show data sections clearly
            if sheet_data['data_sections']:
                print(f"\n   📊 DATA SECTIONS:")
                for i, s in enumerate(sheet_data['data_sections'], 1):
                    print(f"      Section {i}: rows {s['start_row']}-{s['end_row']} ({s['row_count']} rows, ~{s['avg_non_null_per_row']:.0f} filled cols)")

            # Show first few cells to understand structure
            print(f"\n   🔍 FIRST FEW CELLS (row, col, value):")
            for cell in sheet_data['cells'][:10]:
                if cell['value']:
                    print(f"      [{cell['row']}, {cell['col']}] = {cell['value'][:60]}")

            # Column stats summary
            if sheet_data['column_stats']:
                filled_cols = len(sheet_data['column_stats'])
                total_cols = sheet_data['dimensions']['cols']
                print(f"\n   📈 COLUMN USAGE: {filled_cols}/{total_cols} columns have data")
                print(f"      Columns with data: {list(sheet_data['column_stats'].keys())[:10]}")

    def compare_with_schema(self, schema_path: str | Path) -> Dict[str, Any]:
        """
        Compare analysis results with a schema file to identify gaps.
        """
        import yaml

        if not self.analysis:
            raise ValueError("No analysis results. Run analyze() first.")

        schema_path = Path(schema_path)
        with open(schema_path) as f:
            schema = yaml.safe_load(f)

        comparison = {
            "schema_file": str(schema_path),
            "schema_fields": {},
            "coverage_by_sheet": {},
            "gaps": [],
            "duplicates": []
        }

        # Track extraction points
        extraction_points = defaultdict(list)

        # Analyze matrix_fields
        matrix_fields = schema.get("matrix_fields", [])
        for mf in matrix_fields:
            matrix = mf["matrix"]
            field_name = mf["field_name"]
            cell_ref = str(matrix)

            extraction_points[cell_ref].append(f"matrix_field: {field_name}")
            comparison["schema_fields"][field_name] = {
                "type": "matrix_field",
                "location": matrix
            }

        # Analyze named_fields from sheets
        sheets = schema.get("sheets", [])
        for sheet_def in sheets:
            sheet_idx = sheet_def["sheet_index"]
            sheet_type = sheet_def.get("sheet_type")
            named_fields = sheet_def.get("named_fields", [])

            for nf in named_fields:
                row = nf["row"]
                col = nf["column"]
                field_name = nf["field_name"]
                cell_ref = f"[{sheet_idx}, {row}, {col}]"

                extraction_points[cell_ref].append(f"named_field: {field_name}")
                comparison["schema_fields"][field_name] = {
                    "type": "named_field",
                    "location": [sheet_idx, row, col],
                    "sheet_type": sheet_type
                }

        # Check for duplicates
        for cell_ref, extractions in extraction_points.items():
            if len(extractions) > 1:
                comparison["duplicates"].append({
                    "cell": cell_ref,
                    "extracted_as": extractions
                })

        # Calculate coverage per sheet
        for key, sheet_data in self.analysis["sheets"].items():
            if isinstance(sheet_data, SheetAnalysis):
                sheet_data = sheet_data.to_dict()

            if "error" in sheet_data:
                continue

            sheet_idx = sheet_data["sheet_index"]
            total_cells = sheet_data["non_empty_cell_count"]

            # Count extractions from this sheet
            named_from_sheet = sum(1 for f in comparison["schema_fields"].values()
                                  if f.get("type") == "named_field" and
                                  f.get("location", [None])[0] == sheet_idx)

            comparison["coverage_by_sheet"][sheet_data["sheet_name"]] = {
                "sheet_index": sheet_idx,
                "total_non_empty_cells": total_cells,
                "named_field_extractions": named_from_sheet,
                "coverage_pct": round(named_from_sheet / max(total_cells, 1) * 100, 2)
            }

        return comparison


def main():
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Analyze Excel multi-sheet workbooks using pylightxl",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
        """
    )

    parser.add_argument("file", type=str, help="Path to Excel file")
    parser.add_argument("--output", "-o", type=str, help="Output JSON file path")
    parser.add_argument("--sheets", nargs="+", type=int, help="Specific sheet indices to analyze")
    parser.add_argument("--schema", type=str, help="Schema YAML file to compare with")
    parser.add_argument("--max-summary-sheets", type=int, default=None,
                       help="Max sheets to show in summary (default: all)")

    args = parser.parse_args()

    # Create analyzer
    analyzer = ExcelAnalyzer(args.file)

    # Run analysis
    print(f"Analyzing {args.file}...")
    analyzer.analyze(sheets=args.sheets)

    # Print summary
    analyzer.print_summary(max_sheets=args.max_summary_sheets)

    # Save JSON if requested
    if args.output:
        analyzer.save_json(args.output)

    # Compare with schema if provided
    if args.schema:
        print(f"\n{'=' * 100}")
        print("SCHEMA COMPARISON")
        print("=" * 100)
        comparison = analyzer.compare_with_schema(args.schema)

        print(f"\nSchema file: {comparison['schema_file']}")
        print(f"Schema fields defined: {len(comparison['schema_fields'])}")

        if comparison["duplicates"]:
            print(f"\n⚠  DUPLICATES: {len(comparison['duplicates'])}")
            for dup in comparison["duplicates"]:
                print(f"  Cell {dup['cell']}:")
                for e in dup['extracted_as']:
                    print(f"    - {e}")
        else:
            print("\n[OK] No duplicate extractions")

        print(f"\nCoverage by sheet:")
        for sheet_name, cov in comparison["coverage_by_sheet"].items():
            print(f"  {sheet_name:30s} {cov['named_field_extractions']:3d} extractions / {cov['total_non_empty_cells']:5d} cells ({cov['coverage_pct']:5.1f}%)")


if __name__ == "__main__":
    main()
