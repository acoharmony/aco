# © 2025 HarmonyCares
# All rights reserved.

"""
Schema-driven multi-sheet Excel parser with matrix-based table extraction.

WHY THIS EXISTS
===============
CMS REACH reports (PLARU, PALMR, BNMR) contain multiple sheets with different structures:
- Metadata sheet: key-value pairs (2 columns)
- Payment history: standard table with headers
- Detailed calculations: multi-level headers with grouped columns
- Claims data: wide tables with 40+ columns

A single-sheet parser cannot handle this heterogeneity. Each sheet needs different
extraction logic (pivot vs. table vs. multi-level headers), but all must be processed
together to maintain referential integrity (they share metadata like source file, report date).

Without this parser, we'd need separate files for each sheet type, losing the relational
connections between metadata and fact tables.

PROBLEM SOLVED
==============
PLARU.xlsx has 9 sheets:
  Sheet 0: REPORT_PARAMETERS (2 cols, key-value) → needs pivoting
  Sheet 1: PAYMENT_HISTORY (4 cols, standard table) → standard extraction
  Sheet 2: BASE_PCC_PMT_DETAILED (15 cols, multi-level headers) → header combination
  Sheet 3-8: Other calculation sheets with varying structures

This parser:
1. Reads all sheets in ONE pass (efficient)
2. Applies DIFFERENT extraction logic per sheet based on schema config
3. Outputs MULTIPLE tables with shared metadata columns (_output_table partition)
4. Maintains lazy evaluation until final write (Polars-native)

Result: Single LazyFrame with all sheets concatenated, partitioned by _output_table column.

Key Features

- **Multi-Output Support**: Single LazyFrame → multiple output tables via partitioning
- **Per-Sheet Configurations**: Different extraction logic per sheet type
- **Matrix-Based Extraction**: [sheet, row, col] addressing for metadata fields
- **Dynamic Table Detection**: Auto-detect boundaries with end markers
- **Lazy Evaluation**: Stays lazy until write time for memory efficiency
- **Expression Integration**: Chains with multi_level_header, key_value_pivot expressions

Architecture Pattern

Parser outputs single LazyFrame with _output_table column:
  _output_table | other_columns...
  "meta"        | key1, value1, ...
  "meta"        | key2, value2, ...
  "payment"     | date1, amount1, ...
  "payment"     | date2, amount2, ...

Writer splits by _output_table, drops nulls, writes separate parquet files.
This maintains Polars lazy evaluation while enabling multi-table output.

Use Cases

1. **PLARU Reports**: 9 sheets → plaru_meta.parquet + plaru_payment_history.parquet + ...
2. **PALMR Reports**: Payment allocation with multiple calculation sheets
3. **BNMR Reports**: Beneficiary-level data with multiple entity types
4. **PYRED Reports**: 6 service type sheets (inpatient, SNF, home health, etc.)
"""

from pathlib import Path
from typing import Any

import polars as pl
from pydantic import BaseModel, Field

from .._decor8 import parser_method

try:
    from .._trace import TracerWrapper
except ImportError:
    TracerWrapper = None
from ._date_handler import apply_date_parsing
from ._registry import register_parser

# Initialize tracer for this parser
tracer = TracerWrapper("parser.excel_multi_sheet") if TracerWrapper else None


class SheetColumnConfig(BaseModel):
    """Configuration for a single column in a sheet."""

    name: str = Field(..., description="Column name in output")
    position: int = Field(..., description="0-based column index in Excel sheet")
    data_type: str = Field(..., description="Data type: string, integer, decimal, date, boolean")
    description: str | None = Field(None, description="Column description")
    date_format: list[str] | None = Field(
        None, description="Date format patterns if data_type=date"
    )


class MatrixFieldExtraction(BaseModel):
    """
    Configuration for extracting a specific cell value by matrix coordinates.

        Allows extracting metadata or header values from specific locations
        in the workbook, specified as [sheet_index, row_index, column_index].

        Extract report date from row 0, column 14 of all sheets:
            matrix: [null, 0, 14]  # null means apply to all sheets
            field_name: "report_date"
            data_type: "string"
    """

    matrix: list[int | None] = Field(
        ...,
        description="3-element list [sheet_index, row_index, column_index]. Use null for 'all sheets'",
    )
    field_name: str = Field(..., description="Name of field in output")
    data_type: str = Field(default="string", description="Data type of extracted value")
    extract_pattern: str | None = Field(
        None,
        description="Regex pattern to extract portion of cell value (e.g., r'\\d{4}' for year)",
    )
    default_value: Any = Field(
        None, description="Default value if cell is empty or extraction fails"
    )


class SheetConfig(BaseModel):
    """Configuration for parsing a single sheet."""

    sheet_index: int = Field(..., description="0-based sheet index in workbook")
    sheet_type: str = Field(
        ..., description="Sheet type identifier (e.g., 'inpatient', 'physician')"
    )
    description: str | None = Field(None, description="Sheet description")
    columns: list[dict[str, Any]] = Field(..., description="Column definitions for this sheet")


class ExcelMultiSheetConfig(BaseModel):
    """Configuration for multi-sheet Excel parsing."""

    header_row: int = Field(
        ..., description="0-based row index for column headers (hint, will search nearby)"
    )
    data_start_row: int = Field(
        ..., description="0-based row index where data starts (hint, uses header+1)"
    )
    end_marker_column: int = Field(..., description="0-based column index to check for end marker")
    end_marker_value: str = Field(
        ..., description="Value in end_marker_column that signals table end"
    )
    column_mapping_strategy: str = Field(
        default="position",
        description="Strategy for column mapping: 'position' (fixed positions) or 'header_match' (match by header text)",
    )
    header_search_text: str | None = Field(
        None,
        description="Text to search for in first column to identify header row (e.g., 'Provider Type')",
    )


def extract_matrix_fields(
    file_path: Path, matrix_fields: list[dict[str, Any]], sheet_index: int | None = None
) -> dict[str, Any]:
    """
    Extract specific field values using matrix coordinates [sheet, row, column].

        Parameters

        file_path : Path
            Path to Excel file
        matrix_fields : list[dict[str, Any]]
            List of matrix field extraction configs
        sheet_index : int | None
            Current sheet index (if processing specific sheet)

        Returns

        dict[str, Any]
            Dictionary of field_name -> extracted value
            {'performance_year': '2025', 'report_date': 'October 21, 2025'}
    """
    import re

    extracted = {}
    # Cache sheets to avoid re-reading
    _sheet_cache: dict[int, pl.DataFrame] = {}

    def _get_sheet(sheet_idx: int) -> pl.DataFrame | None:
        if sheet_idx not in _sheet_cache:
            try:
                _sheet_cache[sheet_idx] = pl.read_excel(
                    file_path,
                    sheet_id=sheet_idx + 1,
                    read_options={"header_row": None, "skip_rows": 0},
                    infer_schema_length=0,
                )
            except Exception:
                _sheet_cache[sheet_idx] = None
        return _sheet_cache[sheet_idx]

    for field_config in matrix_fields:
        matrix = field_config.get("matrix", [])
        if len(matrix) != 3:
            continue

        target_sheet, target_row, target_col = matrix

        # Check if this field applies to current sheet
        if sheet_index is not None and target_sheet is not None and target_sheet != sheet_index:
            continue

        # Determine which sheet to read from
        read_sheet = target_sheet if target_sheet is not None else sheet_index
        if read_sheet is None:
            continue

        try:
            df = _get_sheet(read_sheet)
            if df is None or len(df) == 0:
                field_name = field_config.get("field_name")
                extracted[field_name] = field_config.get("default_value")
                continue

            # Label-based lookup: search column 0 for the label text
            # and extract the value from target_col on the matched row
            search_label = field_config.get("search_label")
            if search_label:
                target_row = None
                for row_idx in range(len(df)):
                    cell = df.row(row_idx)[0]
                    if cell and search_label.lower() in str(cell).strip().lower():
                        target_row = row_idx
                        break
                if target_row is None:
                    field_name = field_config.get("field_name")
                    extracted[field_name] = field_config.get("default_value")
                    continue

            # Extract value at target row, column
            if target_row < len(df) and target_col < len(df.columns):
                value = df.row(target_row)[target_col]

                # Apply extraction pattern if provided
                extract_pattern = field_config.get("extract_pattern")
                if extract_pattern and value:
                    match = re.search(extract_pattern, str(value))
                    value = (match.group(1) if match.lastindex else match.group(0)) if match else field_config.get("default_value")
                elif value is None:
                    value = field_config.get("default_value")

                # Cast to data type (always cast, even for None values, to maintain schema consistency)
                data_type = field_config.get("data_type", "string")
                if value is not None:
                    if data_type == "integer":
                        value = int(value)
                    elif data_type == "decimal":
                        value = float(value)
                    elif data_type == "string":
                        value = str(value)
                else:
                    # For None values, ensure they're properly typed nulls
                    # by using Polars null casting in the calling code
                    # Store as pl.Utf8 null for strings
                    pass

                field_name = field_config.get("field_name")
                extracted[field_name] = value
            else:
                # Cell out of bounds, use default value
                field_name = field_config.get("field_name")
                extracted[field_name] = field_config.get("default_value")

        except (
            Exception
        ):  # ALLOWED: Excel cell extraction fallback - use default if extraction fails
            # If extraction fails, use default value
            field_name = field_config.get("field_name")
            extracted[field_name] = field_config.get("default_value")

    return extracted


def map_columns_by_position(
    df: pl.DataFrame, columns: list[dict[str, Any]]
) -> tuple[dict[str, str], dict[str, Any]]:
    """
    Map columns by fixed position (original behavior).

        Parameters

        df : pl.DataFrame
            DataFrame with columns to map
        columns : list[dict[str, Any]]
            Column definitions with 'position' field

        Returns

        tuple[dict[str, str], dict[str, Any]]
            (column_mapping from actual to output names, column_dtypes)
    """
    column_mapping = {}
    dtypes = {}
    used_output_names = set()  # Track which output names have been used to prevent duplicates

    for col_def in columns:
        # Try both 'position' and 'source_position'
        position = col_def.get("position")
        if position is None:
            position = col_def.get("source_position")
        if isinstance(position, int) and position < len(df.columns):
            actual_col = df.columns[position]
            output_name = col_def.get("output_name", col_def.get("name"))

            # Skip if this output name has already been mapped (prevent duplicates)
            if output_name in used_output_names:
                continue

            column_mapping[actual_col] = output_name
            used_output_names.add(output_name)

            # Store dtype
            data_type = col_def.get("data_type", "string").lower()
            dtypes[output_name] = data_type

    return column_mapping, dtypes


def map_columns_by_header_match(
    df: pl.DataFrame, header_row_idx: int, columns: list[dict[str, Any]]
) -> tuple[dict[str, str], dict[str, Any], dict[str, dict[str, Any]]]:
    """
    Map columns by matching header text (handles sparse columns with empty cells).

        This strategy finds each column by looking for header text in the header row,
        allowing for empty/merged columns that shift positions between files.

        Also extracts metadata from column headers when specified in column definitions.

        Parameters

        df : pl.DataFrame
            Full dataframe with headers
        header_row_idx : int
            Index of the header row
        columns : list[dict[str, Any]]
            Column definitions with 'header_text' field for matching,
            and optional 'extract_header_metadata' configuration

        Returns

        tuple[dict[str, str], dict[str, Any], dict[str, dict[str, Any]]]
            (column_mapping from actual to output names,
             column_dtypes,
             header_metadata per column: {col_name: {metadata_field: value}})
    """
    column_mapping = {}
    dtypes = {}
    header_metadata = {}  # Store extracted metadata per column
    used_output_names = set()  # Track which output names have been mapped to prevent duplicates

    # Get header row
    header_row = df.row(header_row_idx)

    for col_def in columns:
        # Try header_text first, then source_name, then fall back to name
        header_text = (
            col_def.get("header_text") or col_def.get("source_name") or col_def.get("name")
        )
        output_name = col_def.get("output_name", col_def.get("name"))
        data_type = col_def.get("data_type", "string").lower()

        # Find column index by matching header text
        # Use case-insensitive partial matching with key words
        for col_idx, header_val in enumerate(header_row):
            if not header_val:
                continue

            cell_text = str(header_val).strip().lower()
            schema_text = header_text.lower()

            # Check for partial match in either direction
            # Remove common prefixes/suffixes and normalize whitespace
            import re

            cell_clean = re.sub(
                r"\s+", " ", cell_text.replace("# of ", "")
            )  # Normalize multiple spaces
            schema_clean = re.sub(r"\s+", " ", schema_text.replace("# of ", ""))

            # Match if key words from schema appear in cell text
            # Split on spaces and check if main words match
            # Keep words > 2 chars (to catch "CCN") and not parenthetical notes
            schema_words = [w for w in schema_clean.split() if len(w) > 2 and not w.startswith("(")]

            # For short schema text (like "CCN"), do exact match
            # For longer text, check if first 2 significant words appear
            matched = False
            if len(schema_clean) <= 5:
                # Short header like "CCN" - check exact match or very close match
                if schema_clean == cell_clean or schema_clean in cell_clean.split():
                    matched = True
            elif schema_words and all(
                word in cell_clean for word in schema_words[:2]
            ):  # Match first 2 significant words
                matched = True

            if matched:
                # Skip if this output_name has already been mapped to prevent duplicates
                if output_name in used_output_names:
                    continue

                actual_col = df.columns[col_idx]
                column_mapping[actual_col] = output_name
                dtypes[output_name] = data_type
                used_output_names.add(output_name)

                # Extract metadata from header if specified
                header_metadata_config = col_def.get("extract_header_metadata")
                if header_metadata_config and isinstance(header_metadata_config, list):
                    if output_name not in header_metadata:
                        header_metadata[output_name] = {}

                    for meta_config in header_metadata_config:
                        field_name = meta_config.get("field_name")
                        extract_pattern = meta_config.get("extract_pattern")

                        if field_name and extract_pattern:
                            import re

                            match = re.search(extract_pattern, str(header_val))
                            if match:
                                # Use first capturing group if available, otherwise full match
                                extracted_value = (
                                    match.group(1) if match.lastindex else match.group(0)
                                )
                                header_metadata[output_name][field_name] = extracted_value

                break

    return column_mapping, dtypes, header_metadata


def parse_sheet_matrix(
    file_path: Path,
    sheet_index: int,
    config: ExcelMultiSheetConfig,
    columns: list[dict[str, Any]],
    skip_data_slice: bool = False,
) -> tuple[pl.LazyFrame, dict[str, dict[str, Any]]]:
    """
    Parse a single sheet using matrix-based (row, column) extraction.

        This function reads an Excel sheet and extracts data based on exact
        row and column positions defined in the schema. No type inference
        is performed - all types come from the schema.

        Also extracts metadata from column headers when specified in column definitions.

        Parameters

        file_path : Path
            Path to Excel file
        sheet_index : int
            0-based sheet index to read
        config : ExcelMultiSheetConfig
            Configuration for table boundaries and structure
        columns : list[dict[str, Any]]
            Column definitions with name, position, data_type

        Returns

        tuple[pl.LazyFrame, dict[str, dict[str, Any]]]
            (Extracted data with schema-defined columns and types (as LazyFrame),
             Header metadata extracted from column headers)

        Notes

        The function:
        1. Reads the entire sheet without headers
        2. Finds the data end row by checking end_marker_column
        3. Extracts data rows from data_start_row to end row
        4. Selects columns by position
        5. Renames to schema-defined names
        6. Casts to schema-defined types
    """
    # For sheets without explicit column definitions, use Polars' native header reading
    if not columns and config.header_row is not None:
        # Try reading with polars first
        try:
            import openpyxl

            # Get column headers first to build schema overrides
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                ws = wb.worksheets[sheet_index]
                header_row_data = list(
                    ws.iter_rows(
                        min_row=config.header_row, max_row=config.header_row, values_only=True
                    )
                )
                headers = (
                    [h for h in header_row_data[0] if h is not None] if header_row_data else []
                )
                wb.close()

                # Create schema_overrides to force all columns to String
                schema_overrides = dict.fromkeys(headers, pl.String)
            except Exception:
                # If openpyxl header reading fails (e.g., drawing errors), use no overrides
                schema_overrides = {}

            # Try reading with polars
            df_with_headers = pl.read_excel(
                file_path,
                sheet_id=sheet_index + 1,
                read_options={"header_row": config.header_row - 1, "skip_rows": 0},
                schema_overrides=schema_overrides,
            )

        except (OverflowError, Exception) as e:
            # Fallback to python-calamine (fastexcel) directly if polars fails
            # This bypasses openpyxl's drawing issues and is more modern
            try:
                from fastexcel import read_excel as fastexcel_open

                # Read with fastexcel directly, which uses calamine (Rust-based Excel reader)
                # This avoids openpyxl's drawing XML errors
                reader = fastexcel_open(str(file_path))
                excel_sheet = reader.load_sheet(
                    sheet_index, header_row=config.header_row - 1 if config.header_row else None
                )

                # Convert to polars DataFrame
                df_with_headers = excel_sheet.to_polars()

            except Exception as fastexcel_error:
                # Final fallback to openpyxl
                import openpyxl

                try:
                    # Use openpyxl with minimal options to avoid drawing errors
                    wb = openpyxl.load_workbook(
                        file_path, data_only=True, keep_links=False, keep_vba=False
                    )
                except Exception:
                    # If that still fails, we have to skip drawing loading entirely
                    # by reading the raw worksheet XML
                    raise RuntimeError(
                        f"Cannot read Excel file due to corrupt structure: {fastexcel_error}"
                    ) from e

                ws = wb.worksheets[sheet_index]

                # Extract all rows as strings
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append([str(cell) if cell is not None else None for cell in row])

                wb.close()

                # Create DataFrame from list of lists
                if len(data) > config.header_row:
                    headers = data[config.header_row - 1]
                    rows = data[config.header_row :]

                    # Filter out None headers
                    valid_cols = {}
                    for i, h in enumerate(headers):
                        if h is not None:
                            valid_cols[str(h)] = [
                                row[i] if i < len(row) else None for row in rows
                            ]

                    df_with_headers = pl.DataFrame(valid_cols) if valid_cols else pl.DataFrame()
                else:
                    df_with_headers = pl.DataFrame()

        # Clean column names
        clean_columns = {}
        for col in df_with_headers.columns:
            clean_name = str(col).strip().lower().replace(" ", "_").replace("-", "_")
            clean_columns[col] = clean_name

        df_cleaned = df_with_headers.rename(clean_columns)

        # Apply end_marker filtering if configured
        if config.end_marker_value and config.end_marker_column < len(df_cleaned.columns):
            col_name = df_cleaned.columns[config.end_marker_column]
            # Find first row where end marker appears
            for idx in range(len(df_cleaned)):
                cell_val = df_cleaned[col_name][idx]
                if cell_val and str(cell_val).strip() == config.end_marker_value:
                    df_cleaned = df_cleaned.slice(0, idx)
                    break

        return df_cleaned.lazy(), {}

    # For sheets with explicit column definitions, use matrix extraction
    # Read entire sheet without header inference - force all columns to string
    # Schema drives types, not Excel inference

    # Read entire sheet without header inference - all columns as strings
    # When header_row=None, polars names columns as column_0, column_1, etc.
    # We cannot use schema_overrides with header_row=None because polars
    # doesn't know the column names yet. Instead, read with infer_schema_length=0.
    df_raw = pl.read_excel(
        file_path,
        sheet_id=sheet_index + 1,  # polars uses 1-based sheet IDs
        read_options={"header_row": None, "skip_rows": 0},
        infer_schema_length=0,  # Force all columns to string type
    )

    # Dynamically find header row if search text provided
    header_row_idx = config.header_row
    if config.header_search_text:
        search_start = max(0, config.header_row - 2)
        search_end = min(len(df_raw), config.header_row + 5)

        for row_idx in range(search_start, search_end):
            row = df_raw.row(row_idx)
            first_col = str(row[0]).strip() if row[0] else ""
            if first_col and config.header_search_text in first_col:
                header_row_idx = row_idx
                break

    # Data starts immediately after header
    data_start_row = header_row_idx + 1 if header_row_idx is not None else 0

    # Find table end by checking end_marker_column for end_marker_value
    end_row = None
    for row_idx in range(data_start_row, len(df_raw)):
        row = df_raw.row(row_idx)
        cell_value = row[config.end_marker_column] if config.end_marker_column < len(row) else None
        if cell_value and str(cell_value).strip() == config.end_marker_value:
            end_row = row_idx
            break

    if end_row is None:
        # No end marker found, use all rows
        end_row = len(df_raw)

    # Extract data rows (from data_start_row to end_row, exclusive)
    # For multi-level header transforms, skip slicing to preserve header rows
    if skip_data_slice:
        df_data = df_raw  # Keep all rows including headers
    else:
        df_data = df_raw.slice(data_start_row, end_row - data_start_row)

    # Build column mapping using configured strategy
    header_metadata = {}
    column_mapping = {}
    dtypes = {}

    # Skip column mapping for multi-level header transforms
    # These sheets need raw column_N names so the transform can extract headers
    if skip_data_slice:
        column_mapping = {}  # Keep generic column_N names
        dtypes = {}
    # Only apply column mapping if columns are defined
    # Sheets with only transforms (no columns) will keep generic column_N names
    elif columns:
        # Check if any columns use position-based mapping
        position_columns = [c for c in columns if "source_position" in c or "position" in c]
        header_columns = [c for c in columns if "source_name" in c or "header_text" in c]

        if position_columns and header_columns:
            # Mixed: use both position and header matching
            # First apply position mapping
            column_mapping, dtypes = map_columns_by_position(df_data, position_columns)
            # Then apply header matching for remaining columns
            header_mapping, header_dtypes, header_metadata = map_columns_by_header_match(
                df_raw, header_row_idx, header_columns
            )
            column_mapping.update(header_mapping)
            dtypes.update(header_dtypes)
        elif position_columns:
            # Only position mapping
            column_mapping, dtypes = map_columns_by_position(df_data, position_columns)
        elif header_columns or config.column_mapping_strategy == "header_match":
            # Only header matching
            column_mapping, dtypes, header_metadata = map_columns_by_header_match(
                df_raw, header_row_idx, columns
            )
        else:
            # Fallback to position mapping
            column_mapping, dtypes = map_columns_by_position(df_data, columns)

    # Select and rename columns
    # Always use full mapping (select then rename) to avoid duplicate column issues
    if column_mapping:
        # Full mapping: select only the mapped columns then rename
        df_selected = df_data.select(list(column_mapping.keys()))
        df_selected = df_selected.rename(column_mapping)
    else:
        df_selected = df_data

    # Cast types based on schema
    if column_mapping:
        cast_exprs = []
        for col_def in columns:
            col_name = col_def.get("output_name", col_def.get("name"))
            if col_name not in df_selected.columns:
                continue

            data_type = col_def.get("data_type", "string").lower()

            if data_type in ["int", "integer"]:
                cast_exprs.append(pl.col(col_name).cast(pl.Int64, strict=False))
            elif data_type in ["float", "decimal"]:
                cast_exprs.append(pl.col(col_name).cast(pl.Float64, strict=False))
            elif data_type == "boolean":
                # Handle multiple boolean representations
                cast_exprs.append(
                    pl.when(
                        pl.col(col_name)
                        .cast(pl.Utf8)
                        .is_in(["1", "true", "True", "TRUE", "T", "t", "yes", "Yes", "YES"])
                    )
                    .then(True)
                    .when(
                        pl.col(col_name)
                        .cast(pl.Utf8)
                        .is_in(["0", "false", "False", "FALSE", "F", "f", "no", "No", "NO"])
                    )
                    .then(False)
                    .otherwise(None)
                    .alias(col_name)
                )
            elif data_type == "string":
                cast_exprs.append(pl.col(col_name).cast(pl.Utf8, strict=False))
            elif data_type == "date":
                # Keep as string for now, will be parsed by apply_date_parsing
                cast_exprs.append(pl.col(col_name).cast(pl.Utf8, strict=False))
            else:
                # Unknown type, keep as-is
                cast_exprs.append(pl.col(col_name))

        df_typed = df_selected.select(cast_exprs)
        # Convert to LazyFrame to enable lazy evaluation downstream
        return df_typed.lazy(), header_metadata

    # Convert to LazyFrame to enable lazy evaluation downstream
    return df_selected.lazy(), header_metadata


def _apply_sheet_transform(
    df_sheet: pl.LazyFrame, transform_config: dict[str, Any]
) -> pl.LazyFrame:
    """
    Apply transform to a parsed sheet based on transform configuration.

         idempotent integration between the Excel parser
        and the expression/transform system. It routes to the appropriate expression
        based on the transform type specified in the schema.

        Parameters

        df_sheet : pl.LazyFrame
            Parsed sheet data (lazy)

        transform_config : dict[str, Any]
            Transform configuration from schema with keys:
            - type: str (required) - Transform type: "key_value_pivot", "multi_level_header",
                    "dynamic_meta_detect", "append_detect"
            - Additional keys specific to transform type

        Returns

        pl.LazyFrame
            Transformed sheet data (stays lazy when possible)
    """
    transform_type = transform_config.get("type")

    if not transform_type:
        return df_sheet

    if transform_type == "key_value_pivot":
        from .._transforms._key_value_pivot import KeyValuePivotConfig, KeyValuePivotExpression

        config_params = {k: v for k, v in transform_config.items() if k != "type"}
        pivot_config = KeyValuePivotConfig(**config_params)
        df_collected = df_sheet.collect()
        df_pivoted = KeyValuePivotExpression.build(df_collected, pivot_config)
        return df_pivoted

    elif transform_type == "multi_level_header":
        from .._expressions._multi_level_header import (
            MultiLevelHeaderConfig,
            MultiLevelHeaderExpression,
        )

        config_params = {k: v for k, v in transform_config.items() if k != "type"}
        data_start_row = config_params.pop("data_start_row", None)
        header_config = MultiLevelHeaderConfig(**config_params)
        return MultiLevelHeaderExpression.apply(df_sheet, header_config, data_start_row)

    elif transform_type == "dynamic_meta_detect":
        from .._expressions._dynamic_meta_detect import (
            DynamicMetaConfig,
            DynamicMetaDetectExpression,
        )

        config_params = {k: v for k, v in transform_config.items() if k != "type"}
        data_start_row = config_params.pop("data_start_row", None)
        meta_config = DynamicMetaConfig(**config_params)
        result_lazy, _metadata = DynamicMetaDetectExpression.apply(
            df_sheet, meta_config, data_start_row
        )
        return result_lazy

    else:
        import logging

        logging.warning(f"Unknown transform type '{transform_type}', skipping transform")
        return df_sheet


def _drop_sparse_columns(df: pl.LazyFrame, partition_col: str) -> pl.LazyFrame:
    """
    Drop columns that are entirely null within each partition group.

    When multiple sheet types with different schemas are diagonally concatenated,
    each partition inherits every other partition's columns as all-null. This
    collects the frame, identifies which columns are all-null per partition,
    and drops those that are all-null across ALL partitions (truly unused columns).
    Columns that have data in at least one partition are kept.
    """
    try:
        collected = df.collect()
    except Exception:
        return df

    if partition_col not in collected.columns:
        return collected.lazy()

    # Find columns that have at least some non-null data
    keep = {partition_col}
    for col in collected.columns:
        if collected[col].drop_nulls().len() > 0:
            keep.add(col)

    drop = [c for c in collected.columns if c not in keep]
    if drop:
        collected = collected.drop(drop)

    return collected.lazy()


@register_parser("excel_multi_sheet", metadata={"extensions": [".xlsx"]})
@parser_method(
    threshold=10.0,
    validate_path="file_path",
)
def parse_excel_multi_sheet(
    file_path: Path,
    schema: Any,
    limit: int | None = None,
    sheet_types: list[str] | None = None,
    multi_output: bool = False,
) -> pl.LazyFrame | dict[str, pl.LazyFrame]:
    """
    Parse multi-sheet Excel workbook using schema-driven matrix extraction.

        This parser uses a convention-based plugin system:
        1. Extracts schema name (e.g., "pyred")
        2. Checks for specialized parser at _parsers/_[schema_name].py
        3. If found, delegates to specialized parser (e.g., _pyred.parse_pyred)
        4. Otherwise, uses generic multi-sheet parsing logic

        This allows schema-specific customization without polluting the base parser.

        This parser extends the base Excel parser to handle workbooks with multiple
        sheets, each with potentially different structures. Each sheet is defined
        in the schema with its own column mappings and table boundaries.

        Parameters

        file_path : Path
            Path to Excel workbook
        schema : Any
            Schema object with:
            - file_format.sheet_config: ExcelMultiSheetConfig
            - sheets: List of SheetConfig objects
        limit : int | None, optional
            Max rows per sheet (for testing)
        sheet_types : list[str] | None, optional
            Filter to specific sheet types (e.g., ['inpatient', 'physician'])
        multi_output : bool, optional
            If True, return dict of LazyFrames (one per sheet/type) instead of single combined frame.
            Keys will be '{schema_name}_{sheet_type}' for data sheets and '{schema_name}_meta' for metadata.

        Returns

        pl.LazyFrame | dict[str, pl.LazyFrame]
            If multi_output=False: Combined data from all sheets with 'sheet_type' column added
            If multi_output=True: Dict mapping table names to LazyFrames
        shape: (6, 2)
        ┌────────────┬───────┐
        │ sheet_type ┆ count │
        │ ---        ┆ ---   │
        │ str        ┆ u32   │
        ╞════════════╪═══════╡
        │ inpatient  ┆ 0     │
        │ snf        ┆ 0     │
        │ hh         ┆ 29    │
        │ hospice    ┆ 16    │
        │ outpatient ┆ 0     │
        │ physician  ┆ 219   │
        └────────────┴───────┘

        Parse only specific sheets:
        ['physician', 'hh']
    """
    # Try to delegate to specialized parser based on schema name
    # Convention: _parsers/_[schema_name].py with parse_[schema_name] function
    schema_name = None
    if isinstance(schema, dict):
        schema_name = schema.get("name")
    elif hasattr(schema, "name"):
        schema_name = schema.name

    if schema_name:
        try:
            # Try to import specialized parser module
            import importlib

            specialized_module = importlib.import_module(
                f"._{schema_name}", package="acoharmony._parsers"
            )
            specialized_parser = getattr(specialized_module, f"parse_{schema_name}", None)

            if specialized_parser:
                # Delegate to specialized parser
                return specialized_parser(file_path, schema, limit=limit, sheet_types=sheet_types)
        except (ImportError, AttributeError):
            # No specialized parser found, fall through to generic implementation
            pass

    # Generic multi-sheet parsing logic
    # Get configuration from schema - handle dict, SimpleNamespace, and TableMetadata
    if isinstance(schema, dict):
        file_format = schema.get("file_format", {})
        sheets_list = schema.get("sheets", [])
        schema_name = schema.get("name", "unknown")
    elif hasattr(schema, "file_format"):
        # Handle both dict and object attributes
        file_format = (
            schema.file_format if isinstance(schema.file_format, dict) else vars(schema.file_format)
        )

        # Get sheets - could be None, list, or other iterable
        sheets_list = getattr(schema, "sheets", None)
        if sheets_list is None:
            sheets_list = []
        elif not isinstance(sheets_list, list):
            sheets_list = list(sheets_list)

        schema_name = getattr(schema, "name", "unknown")
    else:
        raise ValueError("Schema must have file_format for multi-sheet parsing")

    if "sheet_config" not in file_format:
        raise ValueError("Schema must have file_format.sheet_config for multi-sheet parsing")

    if not sheets_list:
        raise ValueError("Schema must have 'sheets' list for multi-sheet parsing")

    # Check if schema specifies multi_output mode
    if not multi_output and isinstance(file_format, dict):
        multi_output = file_format.get("multi_output", False)

    # Parse sheet_config into Pydantic model
    sheet_config_dict = file_format["sheet_config"]
    if not isinstance(sheet_config_dict, dict):
        sheet_config_dict = vars(sheet_config_dict)
    config = ExcelMultiSheetConfig(**sheet_config_dict)

    # Extract matrix fields from schema if present
    matrix_fields_list = []
    if isinstance(schema, dict):
        matrix_fields_list = schema.get("matrix_fields", [])
    elif hasattr(schema, "matrix_fields"):
        matrix_fields_raw = getattr(schema, "matrix_fields", None)
        if matrix_fields_raw:
            if isinstance(matrix_fields_raw, list):
                matrix_fields_list = matrix_fields_raw
            else:
                matrix_fields_list = list(matrix_fields_raw)

    # Convert matrix fields to dicts if needed
    matrix_fields_dicts = []
    if matrix_fields_list:
        for mf in matrix_fields_list:
            if isinstance(mf, dict):
                matrix_fields_dicts.append(mf)
            else:
                matrix_fields_dicts.append(vars(mf))

    # Build sheet name → index lookup for name-based matching
    _sheet_name_to_idx: dict[str, int] = {}
    try:
        import openpyxl as _oxl

        _wb = _oxl.load_workbook(file_path, read_only=True, data_only=True)
        for _i, _name in enumerate(_wb.sheetnames):
            _sheet_name_to_idx[_name.upper()] = _i
        _wb.close()
    except Exception:
        pass

    # Pre-extract global matrix fields (those targeting specific sheets)
    # so they can be applied to ALL sheets as workbook-level metadata
    global_matrix_values = {}
    if matrix_fields_dicts:
        global_matrix_values = extract_matrix_fields(
            file_path, matrix_fields_dicts, sheet_index=None
        )

    # Parse all sheets
    all_sheets = []
    sheet_tables = {}  # For multi_output mode: {table_name: LazyFrame}
    metadata_sheets = []  # For multi_output mode: collect metadata sheets

    for sheet_def in sheets_list:
        # Convert to dict if needed
        if not isinstance(sheet_def, dict):
            sheet_def = vars(sheet_def)

        # Check if we should process this sheet
        sheet_type = sheet_def.get("sheet_type")
        if sheet_types and sheet_type not in sheet_types:
            continue

        # Resolve sheet index: prefer sheet_name lookup, fall back to sheet_index
        sheet_name = sheet_def.get("sheet_name")
        sheet_index = sheet_def.get("sheet_index")
        if sheet_name and _sheet_name_to_idx:
            resolved = _sheet_name_to_idx.get(sheet_name.upper())
            if resolved is not None:
                sheet_index = resolved
            else:
                # Sheet name not found in this workbook — skip
                continue

        columns_raw = sheet_def.get("columns", [])

        # Convert columns to dict if needed
        columns = []
        for col in columns_raw:
            if isinstance(col, dict):
                columns.append(col)
            else:
                columns.append(vars(col))

        # Check if sheet has its own sheet_config override
        sheet_specific_config = config
        if "sheet_config" in sheet_def:
            # Merge sheet-specific config with global config
            sheet_config_override = sheet_def["sheet_config"]
            if not isinstance(sheet_config_override, dict):
                sheet_config_override = vars(sheet_config_override)

            # Create merged config dict
            merged_config_dict = dict(vars(config))  # Start with global config
            merged_config_dict.update(sheet_config_override)  # Override with sheet-specific
            sheet_specific_config = ExcelMultiSheetConfig(**merged_config_dict)

        # Check if sheet has multi-level header transform
        has_mlh_transform = False
        if "transform" in sheet_def:
            transform = sheet_def["transform"]
            if isinstance(transform, dict):
                has_mlh_transform = transform.get("type") == "multi_level_header"

        # Parse this sheet
        try:
            df_sheet, col_header_metadata = parse_sheet_matrix(
                file_path,
                sheet_index,
                sheet_specific_config,
                columns,
                skip_data_slice=has_mlh_transform,
            )
        except Exception as e:
            # Sheet doesn't exist in this file (some files have fewer sheets)
            # Skip this sheet and continue processing
            if "no matching sheet found" in str(e):
                continue
            # Re-raise other errors
            raise

        # DISABLED: Transforms should happen in transform stage, not parser stage
        # Parser outputs raw data with _output_table column
        # SchemaTransformer._apply_multi_sheet_transforms() handles sheet-specific transforms
        # if "transform" in sheet_def:
        #     df_sheet = _apply_sheet_transform(df_sheet, sheet_def["transform"])

        # Add sheet_type column
        df_sheet = df_sheet.with_columns(pl.lit(sheet_type).alias("sheet_type"))

        # Add extracted header metadata as columns (if any)
        # Only add if column doesn't already exist to avoid duplicates
        existing_columns = set(df_sheet.columns)
        for _col_name, metadata_dict in col_header_metadata.items():
            for field_name, field_value in metadata_dict.items():
                if field_name not in existing_columns:
                    df_sheet = df_sheet.with_columns(
                        pl.lit(field_value, dtype=pl.Utf8).alias(field_name)
                    )
                    existing_columns.add(field_name)

        # Apply matrix field values to this sheet
        # Merge global (pre-extracted) values with any sheet-specific extractions
        if matrix_fields_dicts:
            # Start with global values (extracted once from target sheets)
            matrix_field_values = dict(global_matrix_values)
            # Override with sheet-specific extractions (fields targeting this sheet)
            sheet_specific = extract_matrix_fields(
                file_path, matrix_fields_dicts, sheet_index=sheet_index
            )
            matrix_field_values.update(sheet_specific)

            # Add each matrix field as a column with explicit type casting
            # Only add if column doesn't already exist to avoid duplicates
            for mf_config in matrix_fields_dicts:
                field_name = mf_config.get("field_name")
                if field_name in existing_columns:
                    continue  # Skip if already exists

                data_type = mf_config.get("data_type", "string")
                field_value = matrix_field_values.get(field_name)

                # Determine Polars dtype
                if data_type == "integer":
                    pl_dtype = pl.Int64
                elif data_type == "decimal":
                    pl_dtype = pl.Float64
                else:  # string
                    pl_dtype = pl.Utf8

                # Add column with explicit type cast
                df_sheet = df_sheet.with_columns(
                    pl.lit(field_value, dtype=pl_dtype).alias(field_name)
                )
                existing_columns.add(field_name)

        # Apply limit if specified
        if limit:
            df_sheet = df_sheet.head(limit)

        # Check if this is a metadata or data sheet
        sheet_description = sheet_def.get("description", "")
        is_metadata = "metadata" in sheet_description.lower() or sheet_type == "metadata"

        if multi_output:
            # Store in separate dict for multi-output mode
            if is_metadata:
                metadata_sheets.append(df_sheet)
            else:
                table_name = f"{schema_name}_{sheet_type}"
                sheet_tables[table_name] = df_sheet
        else:
            # Combine all sheets (original behavior)
            all_sheets.append(df_sheet)

    # Return based on output mode
    if multi_output:
        # Multi-output mode: return SINGLE LazyFrame with _output_table column
        # Use diagonal concat to handle different schemas, stay lazy, partition at write time
        if not sheet_tables and not metadata_sheets:
            raise ValueError(f"No sheets found matching sheet_types={sheet_types}")

        all_output_sheets = []

        # Add metadata sheets with output table marker
        if metadata_sheets:
            for df_meta in metadata_sheets:
                df_meta = df_meta.with_columns(pl.lit(f"{schema_name}_meta").alias("_output_table"))
                all_output_sheets.append(df_meta)

        # Add data sheets with output table marker
        for table_name, df in sheet_tables.items():
            df = df.with_columns(pl.lit(table_name).alias("_output_table"))
            all_output_sheets.append(df)

        # Combine all sheets into single LazyFrame using diagonal (allows different schemas)
        df_lazy = pl.concat(all_output_sheets, how="diagonal")
        df_lazy = apply_date_parsing(df_lazy, schema)

        return df_lazy
    else:
        # Single-output mode: combine all sheets
        if not all_sheets:
            raise ValueError(f"No sheets found matching sheet_types={sheet_types}")

        # Union all sheets - use diagonal to handle different column sets per sheet
        # Each sheet type may have different columns (e.g., physician vs inpatient)
        # pl.concat on LazyFrames returns a LazyFrame - stays lazy!
        df_lazy = pl.concat(all_sheets, how="diagonal_relaxed")

        # Apply date parsing for date columns
        df_lazy = apply_date_parsing(df_lazy, schema)

        return df_lazy
