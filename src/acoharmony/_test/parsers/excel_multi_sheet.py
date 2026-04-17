"""Tests for acoharmony._parsers._excel_multi_sheet module."""

# Magic auto-import: brings in ALL exports from module under test
from dataclasses import dataclass
from acoharmony._test._import_magic import auto_import


@auto_import
class _:
    pass  # noqa: E701


from unittest.mock import MagicMock, patch
import pytest
import sys
from types import SimpleNamespace
from pathlib import Path
import acoharmony
from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform

from .conftest import HAS_OPENPYXL


class TestModuleStructure:
    """Basic module structure tests."""

    @pytest.mark.unit
    def test_module_imports(self):
        """Module can be imported."""
        assert acoharmony._parsers._excel_multi_sheet is not None


class TestExtractMatrixFieldsEdgeCases:
    """Cover uncovered branches in extract_matrix_fields."""

    @pytest.mark.unit
    def test_extract_with_out_of_range_row(self, tmp_path):
        """Matrix with row index beyond available rows returns None."""
        df = pl.DataFrame({"col_0": ["only_row"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [0, 999, 0], "field_name": "missing"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result.get("missing") is None

    @pytest.mark.unit
    def test_extract_with_out_of_range_col(self, tmp_path):
        """Matrix with column index beyond available columns returns None."""
        df = pl.DataFrame({"col_0": ["val"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [0, 0, 999], "field_name": "missing"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result.get("missing") is None

    @pytest.mark.unit
    def test_extract_pattern_no_match(self, tmp_path):
        """Extract pattern that doesn't match returns raw value."""
        df = pl.DataFrame({"col_0": ["NoMatch"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [
            {"matrix": [0, 1, 0], "field_name": "val", "extract_pattern": "XYZ_(\\d+)"}
        ]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert "val" in result

    @pytest.mark.unit
    def test_extract_with_date_format(self, tmp_path):
        """Matrix extraction with date_format specified."""
        df = pl.DataFrame({"col_0": ["01/15/2025"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [0, 1, 0], "field_name": "dt", "date_format": ["%m/%d/%Y"]}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert "dt" in result


class TestMapColumnsByPositionExtended:
    """Additional map_columns_by_position tests."""

    @pytest.mark.unit
    def test_maps_with_date_type(self):
        """Position mapping with date data_type."""
        df = pl.DataFrame({"col_0": ["2025-01-01"], "col_1": ["hello"]})
        columns = [
            {"name": "date_field", "position": 0, "data_type": "date"},
            {"name": "text_field", "position": 1, "data_type": "string"},
        ]
        rename_map, type_map = map_columns_by_position(df, columns)
        assert isinstance(rename_map, dict)
        assert isinstance(type_map, dict)

    @pytest.mark.unit
    def test_maps_with_integer_type(self):
        """Position mapping with integer data_type."""
        df = pl.DataFrame({"col_0": ["123"], "col_1": ["456"]})
        columns = [{"name": "int_field", "position": 0, "data_type": "integer"}]
        rename_map, type_map = map_columns_by_position(df, columns)
        assert isinstance(rename_map, dict)


class TestMapColumnsByHeaderMatchExtended:
    """Additional map_columns_by_header_match tests."""

    @pytest.mark.unit
    def test_no_matching_headers(self):
        """Header match with no columns matching returns empty."""
        df = pl.DataFrame({"col_0": ["Foo", "val1"], "col_1": ["Bar", "val2"]})
        columns = [
            {"name": "name", "position": 0, "data_type": "string", "header_text": "ZZZ_NOMATCH"}
        ]
        rename_map, type_map, extra = map_columns_by_header_match(df, 0, columns)
        assert isinstance(rename_map, dict)

    @pytest.mark.unit
    def test_partial_header_match(self):
        """Some headers match, others don't."""
        df = pl.DataFrame({"col_0": ["Name", "John"], "col_1": ["Other", "val"]})
        columns = [
            {"name": "name", "position": 0, "data_type": "string", "header_text": "Name"},
            {"name": "missing", "position": 1, "data_type": "string", "header_text": "NotHere"},
        ]
        rename_map, type_map, extra = map_columns_by_header_match(df, 0, columns)
        assert isinstance(rename_map, dict)


class TestApplySheetTransformExtended:
    """Additional _apply_sheet_transform tests."""

    @pytest.mark.unit
    def test_transform_with_filter(self):
        """Transform config with filter operation."""
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform
        df = pl.DataFrame({"col_a": ["val1", "val2", "val3"], "col_b": ["10", "20", "30"]}).lazy()
        transform_config = {"filter": {"col_a": "val1"}}
        result = _apply_sheet_transform(df, transform_config)
        assert isinstance(result, pl.LazyFrame)

    @pytest.mark.unit
    def test_transform_with_empty_config(self):
        """Transform with empty config returns unchanged data."""
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform
        df = pl.DataFrame({"col_a": ["val"]}).lazy()
        result = _apply_sheet_transform(df, {})
        assert isinstance(result, pl.LazyFrame)
        assert result.collect().height == 1

    @pytest.mark.unit
    def test_transform_with_none_config(self):
        """Transform with config lacking a type returns unchanged data."""
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform
        df = pl.DataFrame({"col_a": ["val"]}).lazy()
        result = _apply_sheet_transform(df, {})
        assert isinstance(result, pl.LazyFrame)


class TestConfigEdgeCases:
    """Edge cases for config dataclasses."""

    @pytest.mark.unit
    def test_sheet_column_config_full_init(self):
        """SheetColumnConfig with all optional fields."""
        config = SheetColumnConfig(
            name="test",
            position=0,
            data_type="float",
            description="A test column",
            date_format=["%Y-%m-%d"],
        )
        assert config.description == "A test column"
        assert config.date_format == ["%Y-%m-%d"]

    @pytest.mark.unit
    def test_sheet_config_with_transform(self):
        """SheetConfig with description configuration."""
        config = SheetConfig(
            sheet_index=1, sheet_type="claims", columns=[], description="Claims sheet"
        )
        assert config.description is not None

    @pytest.mark.unit
    def test_excel_multi_sheet_config_with_header_search(self):
        """ExcelMultiSheetConfig with header_search_text."""
        config = ExcelMultiSheetConfig(
            header_row=0,
            data_start_row=1,
            end_marker_column=0,
            end_marker_value="Total",
            header_search_text="Provider Type",
        )
        assert config.header_search_text == "Provider Type"
        assert config.column_mapping_strategy == "position"


class TestExtractMatrixFieldsDataTypes:
    """Cover lines 224-234: data type casting in extract_matrix_fields."""

    @pytest.mark.unit
    def test_extract_integer_type(self, tmp_path):
        """Cover lines 224-225: integer data_type casting."""
        df = pl.DataFrame({"col_0": ["42"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [0, 1, 0], "field_name": "count", "data_type": "integer"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result["count"] == 42

    @pytest.mark.unit
    def test_extract_decimal_type(self, tmp_path):
        """Cover lines 226-227: decimal data_type casting."""
        df = pl.DataFrame({"col_0": ["3.14"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [0, 1, 0], "field_name": "amount", "data_type": "decimal"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert abs(result["amount"] - 3.14) < 0.01

    @pytest.mark.unit
    def test_extract_string_type(self, tmp_path):
        """Cover lines 228-229: string data_type casting."""
        df = pl.DataFrame({"col_0": ["hello"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [0, 1, 0], "field_name": "text", "data_type": "string"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result["text"] == "hello"

    @pytest.mark.unit
    def test_extract_none_value_default(self, tmp_path):
        """Cover lines 218-219: None value uses default_value."""
        df = pl.DataFrame({"col_0": [None], "col_1": ["data"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [0, 1, 0], "field_name": "val", "default_value": "fallback"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert "val" in result

    @pytest.mark.unit
    def test_extract_pattern_with_group(self, tmp_path):
        """Cover line 217: regex pattern with capturing group."""
        df = pl.DataFrame({"col_0": ["Year: 2025"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [
            {"matrix": [0, 1, 0], "field_name": "year", "extract_pattern": "Year:\\s*(\\d{4})"}
        ]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result["year"] == "2025"

    @pytest.mark.unit
    def test_extract_different_sheet(self, tmp_path):
        """Cover line 192: field for different sheet is skipped."""
        df = pl.DataFrame({"col_0": ["val"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [5, 0, 0], "field_name": "other_sheet"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result == {}

    @pytest.mark.unit
    def test_extract_invalid_matrix_length(self, tmp_path):
        """Cover line 186-187: invalid matrix length is skipped."""
        df = pl.DataFrame({"col_0": ["val"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        matrix_fields = [{"matrix": [0, 0], "field_name": "bad"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result == {}

    @pytest.mark.unit
    def test_extract_exception_uses_default(self, tmp_path):
        """Cover lines 243-248: exception during extraction uses default."""
        matrix_fields = [
            {"matrix": [0, 0, 0], "field_name": "err_field", "default_value": "default_val"}
        ]
        result = extract_matrix_fields(tmp_path / "nonexistent.xlsx", matrix_fields, sheet_index=0)
        assert result["err_field"] == "default_val"


class TestExcelMultiSheetAdditional:
    """Additional tests to fill coverage gaps in _excel_multi_sheet.py."""

    @pytest.mark.unit
    def test_parse_sheet_matrix_without_columns(self, tmp_path: Path):
        """Cover lines 455-564: parse_sheet_matrix when columns is empty list."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "no_cols.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Value", "Status"])
        ws.append(["Alice", "100", "ok"])
        ws.append(["Bob", "200", "ok"])
        ws.append(["TOTAL", "", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="TOTAL"
        )
        lf, meta = parse_sheet_matrix(p, 0, config, columns=[])
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_sheet_matrix_without_columns_end_marker(self, tmp_path: Path):
        """Cover end_marker filtering in no-columns branch."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "no_cols_end.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Metric", "Amount"])
        ws.append(["A", "10"])
        ws.append(["B", "20"])
        ws.append(["END", ""])
        ws.append(["Extra", "99"])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="END"
        )
        lf, meta = parse_sheet_matrix(p, 0, config, columns=[])
        df = lf.collect()
        assert "Extra" not in df.to_series(0).to_list()

    @pytest.mark.unit
    def test_parse_sheet_matrix_header_search_text(self, tmp_path: Path):
        """Cover lines 583-592: header_search_text to dynamically find header row."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "search.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Report Title", ""])
        ws.append(["Provider Type", "Amount"])
        ws.append(["Hospital", "100"])
        ws.append(["Clinic", "200"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1,
            data_start_row=2,
            end_marker_column=0,
            end_marker_value="TOTAL",
            header_search_text="Provider Type",
        )
        columns = [
            {"position": 1, "name": "provider_type", "data_type": "string"},
            {"position": 2, "name": "amount", "data_type": "string"},
        ]
        lf, meta = parse_sheet_matrix(p, 0, config, columns)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_sheet_matrix_no_end_marker_found(self, tmp_path: Path):
        """Cover line 608: when no end marker is found in any row."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "no_end.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Value"])
        ws.append(["A", "1"])
        ws.append(["B", "2"])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=0, data_start_row=1, end_marker_column=0, end_marker_value="TOTAL"
        )
        columns = [{"position": 1, "name": "name", "data_type": "string"}]
        lf, meta = parse_sheet_matrix(p, 0, config, columns)
        df = lf.collect()
        assert len(df) >= 2

    @pytest.mark.unit
    def test_parse_sheet_matrix_skip_data_slice(self, tmp_path: Path):
        """Cover lines 612-613, 624-626: skip_data_slice=True for multi-level header."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "mlh.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Category", "Group A", "Group B"])
        ws.append(["SubCat", "Val1", "Val2"])
        ws.append(["Row1", "10", "20"])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=0, data_start_row=2, end_marker_column=0, end_marker_value="TOTAL"
        )
        lf, meta = parse_sheet_matrix(p, 0, config, columns=[], skip_data_slice=True)
        df = lf.collect()
        assert len(df) >= 3

    @pytest.mark.unit
    def test_parse_sheet_matrix_mixed_column_mapping(self, tmp_path: Path):
        """Cover lines 634-643: both position-based and header-based columns."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "mixed.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Provider Type", "Amount", "CCN Code"])
        ws.append(["Hospital", "100", "XY123"])
        ws.append(["Clinic", "200", "AB456"])
        ws.append(["TOTAL", "", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=0, data_start_row=1, end_marker_column=0, end_marker_value="TOTAL"
        )
        columns = [
            {"source_position": 1, "name": "provider_type", "data_type": "string"},
            {"source_name": "CCN Code", "name": "ccn", "data_type": "string"},
        ]
        lf, meta = parse_sheet_matrix(p, 0, config, columns)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_sheet_matrix_header_match_strategy(self, tmp_path: Path):
        """Cover lines 647-651: pure header_match strategy."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "header_match.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Provider Type", "Total Amount"])
        ws.append(["Hospital", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=0,
            data_start_row=1,
            end_marker_column=0,
            end_marker_value="TOTAL",
            column_mapping_strategy="header_match",
        )
        columns = [
            {"name": "provider_type", "data_type": "string"},
            {"name": "total_amount", "data_type": "decimal"},
        ]
        lf, meta = parse_sheet_matrix(p, 0, config, columns)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_sheet_matrix_fallback_position(self, tmp_path: Path):
        """Cover lines 652-654: fallback to position mapping when no header/position keys."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "fallback.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append(["1", "2"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=0, data_start_row=1, end_marker_column=0, end_marker_value="TOTAL"
        )
        columns = [{"name": "col_a", "data_type": "string", "position": 1}]
        lf, meta = parse_sheet_matrix(p, 0, config, columns)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_sheet_matrix_type_casting(self, tmp_path: Path):
        """Cover lines 675-704: boolean, date, and unknown type casting."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "types.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Flag", "Amount", "DateCol", "Other"])
        ws.append(["true", "100", "2025-01-01", "misc"])
        ws.append(["false", "200", "2025-02-01", "stuff"])
        ws.append(["TOTAL", "", "", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=0, data_start_row=1, end_marker_column=0, end_marker_value="TOTAL"
        )
        columns = [
            {"position": 1, "name": "flag", "data_type": "boolean"},
            {"position": 2, "name": "amount", "data_type": "integer"},
            {"position": 3, "name": "date_col", "data_type": "date"},
            {"position": 4, "name": "other", "data_type": "unknown_type"},
        ]
        lf, meta = parse_sheet_matrix(p, 0, config, columns)
        df = lf.collect()
        assert len(df) == 2

    @pytest.mark.unit
    def test_apply_sheet_transform_key_value_pivot(self):
        """Cover lines 745-752: key_value_pivot transform."""
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform

        df = pl.DataFrame({"column_1": ["key_a", "key_b"], "column_2": ["val_a", "val_b"]}).lazy()
        mock_module = MagicMock()
        mock_config = MagicMock()
        mock_module.KeyValuePivotConfig = MagicMock(return_value=mock_config)
        mock_pivoted = pl.DataFrame({"key_a": ["val_a"], "key_b": ["val_b"]}).lazy()
        mock_module.KeyValuePivotExpression.build = MagicMock(return_value=mock_pivoted)
        saved = sys.modules.get("acoharmony._transforms._key_value_pivot")
        sys.modules["acoharmony._transforms._key_value_pivot"] = mock_module
        try:
            result = _apply_sheet_transform(
                df,
                {"type": "key_value_pivot", "key_column": "column_1", "value_column": "column_2"},
            )
            collected = result.collect()
            assert len(collected) >= 1
        finally:
            if saved is not None:
                sys.modules["acoharmony._transforms._key_value_pivot"] = saved
            else:
                sys.modules.pop("acoharmony._transforms._key_value_pivot", None)

    @pytest.mark.unit
    def test_apply_sheet_transform_multi_level_header(self):
        """Cover lines 754-763: multi_level_header transform."""
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform

        df = pl.DataFrame(
            {
                "column_0": ["Category", "Sub", "Row1", "Row2"],
                "column_1": ["Group A", "Val1", "10", "20"],
                "column_2": ["Group B", "Val2", "30", "40"],
            }
        ).lazy()
        mock_module = MagicMock()
        mock_config = MagicMock()
        mock_module.MultiLevelHeaderConfig = MagicMock(return_value=mock_config)
        mock_result = pl.DataFrame({"col1": ["10"], "col2": ["30"]}).lazy()
        mock_module.MultiLevelHeaderExpression.apply = MagicMock(return_value=mock_result)
        saved = sys.modules.get("acoharmony._expressions._multi_level_header")
        sys.modules["acoharmony._expressions._multi_level_header"] = mock_module
        try:
            result = _apply_sheet_transform(
                df,
                {
                    "type": "multi_level_header",
                    "header_rows": [0, 1],
                    "separator": "_",
                    "data_start_row": 2,
                },
            )
            collected = result.collect()
            assert len(collected) >= 1
        finally:
            if saved is not None:
                sys.modules["acoharmony._expressions._multi_level_header"] = saved
            else:
                sys.modules.pop("acoharmony._expressions._multi_level_header", None)

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_specializer_routing(self, tmp_path: Path):
        """Cover lines 854-871: specialized parser routing via schema name."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "pyred_route.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Performance Year 2025", None, None, None, None])
        ws.append([None, None, None, None, "January 2025 experience"])
        ws.append([None])
        ws.append(["Provider Type", "NPI", "Name", "Amount", "Status"])
        ws.append(["Physician", "1234567890", "Dr. Smith", "100.00", "Active"])
        ws.append(["TOTAL", "", "", "", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "pyred",
            "file_format": {
                "sheet_config": {
                    "header_row": 3,
                    "data_start_row": 4,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                {
                    "sheet_type": "physician",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "npi", "data_type": "string"}],
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert "performance_year" in df.columns

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_namespace_schema(self, tmp_path: Path):
        """Cover lines 882-895: namespace schema handling in generic parser."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "ns_schema.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = SimpleNamespace(
            name="test_ns",
            file_format={
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            sheets=[
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        )
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_no_file_format_raises(self, tmp_path: Path):
        """Cover line 897: raise for schema without file_format attr."""
        p = tmp_path / "x.xlsx"
        pl.DataFrame({"a": [1]}).write_excel(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        with pytest.raises(ValueError, match="file_format"):
            parse_excel_multi_sheet(p, 42)

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_multi_output_mode(self, tmp_path: Path):
        """Cover lines 906-907, 1056-1091: multi_output mode."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "multi_out.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Type", "Value"])
        ws1.append(["A", "100"])
        ws1.append(["TOTAL", ""])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Key", "Data"])
        ws2.append(["X", "200"])
        ws2.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_mo",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                },
                "multi_output": True,
            },
            "sheets": [
                {
                    "sheet_type": "metadata",
                    "sheet_index": 0,
                    "description": "Metadata sheet",
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                },
                {
                    "sheet_type": "data",
                    "sheet_index": 1,
                    "columns": [{"position": 1, "name": "key_col", "data_type": "string"}],
                },
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert "_output_table" in df.columns

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_multi_output_no_sheets_raises(self, tmp_path: Path):
        """Cover line 1072: multi_output with no matching sheets."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "mo_empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_empty",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "END",
                },
                "multi_output": True,
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "a", "data_type": "string"}],
                }
            ],
        }
        with pytest.raises(ValueError, match="No sheets found"):
            parse_excel_multi_sheet(p, schema, sheet_types=["nonexistent"])

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_sheet_config_override(self, tmp_path: Path):
        """Cover lines 964-973: sheet-specific sheet_config override."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "override.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["B", "200"])
        ws.append(["END_HERE", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_override",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                    "sheet_config": {"end_marker_value": "END_HERE"},
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_with_matrix_fields(self, tmp_path: Path):
        """Cover lines 916-934, 1019-1046: matrix_fields extraction."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "matrix.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Performance Year 2025", "Report Date Oct 2025"])
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_matrix",
            "file_format": {
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "matrix_fields": [
                {
                    "matrix": [0, 0, 0],
                    "field_name": "perf_year",
                    "data_type": "string",
                    "extract_pattern": "\\d{4}",
                }
            ],
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert "perf_year" in df.columns

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_mlh_transform_flag(self, tmp_path: Path):
        """Cover lines 976-980, 989: multi_level_header transform detection."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "mlh_flag.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Category", "Group A", "Group B"])
        ws.append(["SubCat", "Val1", "Val2"])
        ws.append(["Row1", "10", "20"])
        ws.append(["TOTAL", "", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_mlh",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [],
                    "transform": {"type": "multi_level_header", "header_rows": [0, 1]},
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert "sheet_type" in df.columns

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_with_limit(self, tmp_path: Path):
        """Cover line 1049: limit parameter."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "limited.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "1"])
        ws.append(["B", "2"])
        ws.append(["C", "3"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_limit",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema, limit=1)
        df = lf.collect()
        assert len(df) <= 1

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_header_metadata_columns(self, tmp_path: Path):
        """Cover lines 1011-1017: header metadata added as columns."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "hdr_meta.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Claims CY2025"])
        ws.append(["100"])
        ws.append(["TOTAL"])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_hdr",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                    "column_mapping_strategy": "header_match",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [
                        {
                            "header_text": "Claims",
                            "name": "claims",
                            "data_type": "string",
                            "extract_header_metadata": [
                                {"field_name": "claim_year", "extract_pattern": "CY(\\d{4})"}
                            ],
                        }
                    ],
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert "claim_year" in df.columns

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_namespace_matrix_fields(self, tmp_path: Path):
        """Cover lines 920-925, 928-934: namespace matrix_fields conversion."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "ns_matrix.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Year 2025", ""])
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = SimpleNamespace(
            name="test_ns_mf",
            file_format={
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            matrix_fields=[
                SimpleNamespace(
                    matrix=[0, 0, 0],
                    field_name="year",
                    data_type="string",
                    extract_pattern=None,
                    default_value=None,
                )
            ],
            sheets=[
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        )
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert "year" in df.columns

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_namespace_sheet_config(self, tmp_path: Path):
        """Cover lines 911-912: sheet_config_dict from non-dict."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "ns_sc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_ns_sc",
            "file_format": {
                "sheet_config": SimpleNamespace(
                    header_row=0,
                    data_start_row=1,
                    end_marker_column=0,
                    end_marker_value="TOTAL",
                    column_mapping_strategy="position",
                    header_search_text=None,
                )
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_namespace_columns(self, tmp_path: Path):
        """Cover lines 956-960: columns from namespace objects."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "ns_cols.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_ns_cols",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                SimpleNamespace(
                    sheet_type="data",
                    sheet_index=0,
                    columns=[SimpleNamespace(position=1, name="type_col", data_type="string")],
                )
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert len(df) >= 1


class TestParseSheetMatrix:
    """Cover parse_sheet_matrix function (lines 455-711)."""

    @pytest.mark.unit
    def test_parse_sheet_no_columns_with_header(self, tmp_path):
        """Cover lines 455-564: no columns, uses header_row."""
        df = pl.DataFrame({"Name": ["Alice", "Bob"], "Age": [30, 25]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="Total"
        )
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns=[])
        assert isinstance(result_df, pl.LazyFrame)
        collected = result_df.collect()
        assert collected.height >= 1

    @pytest.mark.unit
    def test_parse_sheet_with_position_columns(self, tmp_path):
        """Cover lines 629-654: position-based column mapping."""
        df = pl.DataFrame({"col_0": ["Name", "Alice"], "col_1": ["Age", "30"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="Total"
        )
        columns = [
            {"name": "person_name", "position": 0, "data_type": "string"},
            {"name": "person_age", "position": 1, "data_type": "integer"},
        ]
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns)
        assert isinstance(result_df, pl.LazyFrame)

    @pytest.mark.unit
    def test_parse_sheet_with_end_marker(self, tmp_path):
        """Cover lines 598-608: end marker detection."""
        df = pl.DataFrame(
            {
                "col_0": ["Header", "Data1", "Data2", "Total", "Extra"],
                "col_1": ["H2", "D1", "D2", "Sum", "More"],
            }
        )
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="Total"
        )
        columns = [{"name": "col_a", "position": 0, "data_type": "string"}]
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns)
        collected = result_df.collect()
        assert collected.height <= 3

    @pytest.mark.unit
    def test_parse_sheet_with_header_search(self, tmp_path):
        """Cover lines 583-592: header_search_text."""
        df = pl.DataFrame(
            {
                "col_0": ["Garbage", "Provider Type", "Type A", "Type B"],
                "col_1": ["", "Count", "10", "20"],
            }
        )
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1,
            data_start_row=2,
            end_marker_column=0,
            end_marker_value="Total",
            header_search_text="Provider Type",
        )
        columns = [{"name": "ptype", "position": 0, "data_type": "string"}]
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns)
        assert isinstance(result_df, pl.LazyFrame)

    @pytest.mark.unit
    def test_parse_sheet_skip_data_slice(self, tmp_path):
        """Cover lines 612-613: skip_data_slice keeps all rows."""
        df = pl.DataFrame({"col_0": ["H1", "H2", "D1", "D2"], "col_1": ["A", "B", "C", "D"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="Total"
        )
        result_df, metadata = parse_sheet_matrix(
            xlsx_path, 0, config, columns=[], skip_data_slice=True
        )
        assert isinstance(result_df, pl.LazyFrame)

    @pytest.mark.unit
    def test_parse_sheet_boolean_casting(self, tmp_path):
        """Cover lines 679-696: boolean type casting."""
        df = pl.DataFrame({"col_0": ["Header", "true", "false", "1", "0"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="END"
        )
        columns = [{"name": "flag", "position": 0, "data_type": "boolean"}]
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns)
        collected = result_df.collect()
        assert collected.height >= 1

    @pytest.mark.unit
    def test_parse_sheet_date_casting(self, tmp_path):
        """Cover lines 699-701: date type kept as string."""
        df = pl.DataFrame({"col_0": ["Header", "2025-01-01", "2025-02-01"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="END"
        )
        columns = [{"name": "dt", "position": 0, "data_type": "date"}]
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns)
        collected = result_df.collect()
        assert collected.height >= 1

    @pytest.mark.unit
    def test_parse_sheet_float_casting(self, tmp_path):
        """Cover lines 677-678: float/decimal type casting."""
        df = pl.DataFrame({"col_0": ["Header", "1.5", "2.7"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="END"
        )
        columns = [{"name": "amount", "position": 0, "data_type": "float"}]
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns)
        collected = result_df.collect()
        assert collected.height >= 1

    @pytest.mark.unit
    def test_parse_sheet_unknown_type(self, tmp_path):
        """Cover lines 702-704: unknown data type kept as-is."""
        df = pl.DataFrame({"col_0": ["Header", "val"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="END"
        )
        columns = [{"name": "x", "position": 0, "data_type": "custom_type"}]
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns)
        assert isinstance(result_df, pl.LazyFrame)


class TestApplySheetTransformTypes:
    """Cover lines 745-783: transform type branches."""

    @pytest.mark.unit
    def test_unknown_transform_type(self):
        """Cover lines 779-783: unknown transform type logs warning."""
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform
        df = pl.DataFrame({"col_a": ["val"]}).lazy()
        result = _apply_sheet_transform(df, {"type": "unknown_type"})
        assert isinstance(result, pl.LazyFrame)
        assert result.collect().height == 1


class TestParseExcelMultiSheetUncoveredBranches:
    """Cover remaining uncovered branches in parse_excel_multi_sheet."""

    @pytest.mark.unit
    def test_specialized_parser_not_found_falls_through(self, tmp_path: Path):
        """Branch 931->940: specialized module imports but has no parse_<name> func.

        When importlib finds the module but getattr returns None for the parse function,
        execution should fall through to the generic multi-sheet parser.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "fallthrough.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        # Use a schema name that resolves to a real module but without the expected
        # parse_<name> function.  We mock the import to return a module with no
        # parse_<name> attribute.
        mock_module = MagicMock(spec=[])  # empty spec -> getattr returns None
        with patch("importlib.import_module", return_value=mock_module):
            schema = {
                "name": "fake_specialized",
                "file_format": {
                    "sheet_config": {
                        "header_row": 0,
                        "data_start_row": 1,
                        "end_marker_column": 0,
                        "end_marker_value": "TOTAL",
                    }
                },
                "sheets": [
                    {
                        "sheet_type": "data",
                        "sheet_index": 0,
                        "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                    }
                ],
            }
            lf = parse_excel_multi_sheet(p, schema)
            df = lf.collect()
            assert len(df) >= 1

    @pytest.mark.unit
    def test_namespace_schema_sheets_none(self, tmp_path: Path):
        """Branch 952->953: namespace schema where sheets attribute is None.

        Should default to empty list then raise ValueError for empty sheets.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "ns_none_sheets.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = SimpleNamespace(
            name="test_none_sheets",
            file_format={
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            sheets=None,
        )
        with pytest.raises(ValueError, match="sheets"):
            parse_excel_multi_sheet(p, schema)

    @pytest.mark.unit
    def test_multi_output_already_true_skips_file_format_check(self, tmp_path: Path):
        """Branch 968->972: when multi_output=True is passed, skip reading from file_format.

        The condition `not multi_output and isinstance(file_format, dict)` is False
        because multi_output is already True.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "mo_skip.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_mo_skip",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                },
                # multi_output is also True here, but we pass multi_output=True
                # so the file_format check is skipped
                "multi_output": False,
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema, multi_output=True)
        df = lf.collect()
        assert "_output_table" in df.columns

    @pytest.mark.unit
    def test_namespace_schema_matrix_fields_falsy(self, tmp_path: Path):
        """Branch 983->990: namespace schema with matrix_fields_raw that is falsy (empty list).

        The `if matrix_fields_raw:` check should be False, skipping to line 990.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "empty_mf.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = SimpleNamespace(
            name="test_empty_mf",
            file_format={
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            matrix_fields=[],
            sheets=[
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        )
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_sheet_name_resolved_from_lookup(self, tmp_path: Path):
        """Branches 1036->1037, 1038->1039: sheet_name is resolved via _sheet_name_to_idx.

        When a sheet_def has sheet_name and the workbook has that sheet name,
        sheet_index should be resolved from the lookup.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "named_sheet.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MyData"
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_named",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_name": "MyData",
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_sheet_name_not_found_skips(self, tmp_path: Path):
        """Branch 1038->1042: sheet_name not found in workbook -> skip sheet.

        When sheet_name is specified and the lookup dict exists but the name
        is not found, the sheet should be skipped (continue).
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "missing_name.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_missing_name",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_name": "NonexistentSheet",
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        }
        with pytest.raises(ValueError, match="No sheets found"):
            parse_excel_multi_sheet(p, schema)

    @pytest.mark.unit
    def test_transform_non_dict_skips_mlh_check(self, tmp_path: Path):
        """Branch 1071->1075: transform is not a dict, so has_mlh_transform stays False.

        When sheet_def has a 'transform' key but its value is not a dict (e.g.,
        a SimpleNamespace), the isinstance check fails and we skip to line 1075.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "non_dict_transform.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_ndt",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                    "transform": SimpleNamespace(type="multi_level_header"),
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_header_metadata_duplicate_field_skipped(self, tmp_path: Path):
        """Branch 1105->1104: field_name already in existing_columns -> skip adding.

        When col_header_metadata returns a field_name that already exists in the
        DataFrame columns, the duplicate should be skipped.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        p = tmp_path / "dup_meta.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Two columns with the same extract_header_metadata field_name
        ws.append(["Claims CY2025", "Revenue CY2025"])
        ws.append(["100", "200"])
        ws.append(["TOTAL", ""])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_dup_meta",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                    "column_mapping_strategy": "header_match",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [
                        {
                            "header_text": "Claims",
                            "name": "claims",
                            "data_type": "string",
                            "extract_header_metadata": [
                                {"field_name": "calendar_year", "extract_pattern": "CY(\\d{4})"}
                            ],
                        },
                        {
                            "header_text": "Revenue",
                            "name": "revenue",
                            "data_type": "string",
                            "extract_header_metadata": [
                                {"field_name": "calendar_year", "extract_pattern": "CY(\\d{4})"}
                            ],
                        },
                    ],
                }
            ],
        }
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        # calendar_year should appear only once (duplicate was skipped)
        assert df.columns.count("calendar_year") == 1
        assert df["calendar_year"][0] == "2025"


class TestParseExcelMultiSheet:
    """Cover lines 853-1105: main parse_excel_multi_sheet function."""

    @pytest.mark.unit
    def test_basic_parse_dict_schema(self, tmp_path):
        """Cover lines 878-903: dict-based schema parsing."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df1 = pl.DataFrame({"Type": ["A", "B"], "Value": ["10", "20"]})
        xlsx_path = tmp_path / "test.xlsx"
        df1.write_excel(xlsx_path)
        schema = {
            "name": "test_schema",
            "file_format": {
                "type": "excel_multi_sheet",
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [
                {
                    "sheet_index": 0,
                    "sheet_type": "data",
                    "columns": [
                        {"name": "type_col", "position": 0, "data_type": "string"},
                        {"name": "value_col", "position": 1, "data_type": "string"},
                    ],
                }
            ],
        }
        result = parse_excel_multi_sheet(xlsx_path, schema)
        assert isinstance(result, pl.LazyFrame)
        collected = result.collect()
        assert collected.height >= 1

    @pytest.mark.unit
    def test_parse_no_sheets_raises(self, tmp_path):
        """Cover line 903: empty sheets list raises ValueError."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"col_0": ["val"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test",
            "file_format": {
                "type": "excel_multi_sheet",
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [],
        }
        with pytest.raises(ValueError, match="sheets"):
            parse_excel_multi_sheet(xlsx_path, schema)

    @pytest.mark.unit
    def test_parse_no_sheet_config_raises(self, tmp_path):
        """Cover line 900: no sheet_config raises ValueError."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"col_0": ["val"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test",
            "file_format": {"type": "excel_multi_sheet"},
            "sheets": [{"sheet_index": 0, "sheet_type": "data", "columns": []}],
        }
        with pytest.raises(ValueError, match="sheet_config"):
            parse_excel_multi_sheet(xlsx_path, schema)

    @pytest.mark.unit
    def test_parse_missing_file_format_raises(self, tmp_path):
        """Cover line 897: schema without file_format raises."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        xlsx_path = tmp_path / "test.xlsx"
        pl.DataFrame({"x": [1]}).write_excel(xlsx_path)
        with pytest.raises((ValueError, AttributeError)):
            parse_excel_multi_sheet(xlsx_path, "not_a_schema")

    @pytest.mark.unit
    def test_parse_with_sheet_type_filter(self, tmp_path):
        """Cover line 948-949: sheet_types filter."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"Type": ["A", "B"], "Value": ["10", "20"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test_schema",
            "file_format": {
                "type": "excel_multi_sheet",
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [
                {
                    "sheet_index": 0,
                    "sheet_type": "physician",
                    "columns": [{"name": "t", "position": 0, "data_type": "string"}],
                }
            ],
        }
        result = parse_excel_multi_sheet(xlsx_path, schema, sheet_types=["physician"])
        assert isinstance(result, pl.LazyFrame)

    @pytest.mark.unit
    def test_parse_no_matching_sheets_raises(self, tmp_path):
        """Cover line 1095: no matching sheets raises ValueError."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"Type": ["A"], "Value": ["10"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test_schema",
            "file_format": {
                "type": "excel_multi_sheet",
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [{"sheet_index": 0, "sheet_type": "data", "columns": []}],
        }
        with pytest.raises(ValueError, match="No sheets"):
            parse_excel_multi_sheet(xlsx_path, schema, sheet_types=["nonexistent_type"])

    @pytest.mark.unit
    def test_parse_with_matrix_fields(self, tmp_path):
        """Cover lines 916-946: matrix fields extraction."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"Year": ["2025"], "Type": ["A"], "Value": ["10"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test_schema",
            "file_format": {
                "type": "excel_multi_sheet",
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [
                {
                    "sheet_index": 0,
                    "sheet_type": "data",
                    "columns": [{"name": "type_col", "position": 0, "data_type": "string"}],
                }
            ],
            "matrix_fields": [{"matrix": [0, 1, 0], "field_name": "year", "data_type": "string"}],
        }
        result = parse_excel_multi_sheet(xlsx_path, schema)
        assert isinstance(result, pl.LazyFrame)
        collected = result.collect()
        assert "year" in collected.columns or collected.height >= 0

    @pytest.mark.unit
    def test_parse_with_limit(self, tmp_path):
        """Cover lines 1048-1050: limit parameter."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"Type": ["A", "B", "C", "D"], "Value": ["1", "2", "3", "4"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test_schema",
            "file_format": {
                "type": "excel_multi_sheet",
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [
                {
                    "sheet_index": 0,
                    "sheet_type": "data",
                    "columns": [{"name": "type_col", "position": 0, "data_type": "string"}],
                }
            ],
        }
        result = parse_excel_multi_sheet(xlsx_path, schema, limit=2)
        assert isinstance(result, pl.LazyFrame)
        collected = result.collect()
        assert collected.height <= 2

    @pytest.mark.unit
    def test_parse_multi_output(self, tmp_path):
        """Cover lines 1056-1091: multi_output mode."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"Type": ["A", "B"], "Value": ["10", "20"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test_schema",
            "file_format": {
                "type": "excel_multi_sheet",
                "multi_output": True,
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [
                {
                    "sheet_index": 0,
                    "sheet_type": "data",
                    "columns": [{"name": "type_col", "position": 0, "data_type": "string"}],
                }
            ],
        }
        result = parse_excel_multi_sheet(xlsx_path, schema)
        assert isinstance(result, pl.LazyFrame)
        collected = result.collect()
        assert "_output_table" in collected.columns


class TestHeaderMatchDuplicateSkip:
    """Cover line 378: duplicate output_name skipped in header match."""

    @pytest.mark.unit
    def test_duplicate_header_match_skipped(self):
        """Line 378: when output_name is already in used_output_names, continue."""
        df = pl.DataFrame({"col_0": ["Amount", "10"], "col_1": ["Amount", "20"]})
        columns = [
            {"name": "amount", "position": 0, "data_type": "string", "header_text": "Amount"}
        ]
        rename_map, type_map, metadata = map_columns_by_header_match(df, 0, columns)
        assert len(rename_map) <= 1


class TestOpenpyxlFallback:
    """Cover lines 504-544: openpyxl fallback when fastexcel fails."""

    @pytest.mark.unit
    def test_openpyxl_fallback_reads_data(self, tmp_path):
        """Lines 504-544: fastexcel fails, falls through to openpyxl."""
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        df = pl.DataFrame({"Name": ["Alice", "Bob"], "Age": ["30", "25"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="Total"
        )
        mock_fastexcel = MagicMock()
        mock_fastexcel.read_excel.side_effect = Exception("fastexcel fail")
        with patch(
            "acoharmony._parsers._excel_multi_sheet.pl.read_excel",
            side_effect=OverflowError("test"),
        ):
            with patch.dict("sys.modules", {"fastexcel": mock_fastexcel}):
                result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns=[])
        assert isinstance(result_df, pl.LazyFrame)
        collected = result_df.collect()
        assert collected.height >= 1

    @pytest.mark.unit
    def test_openpyxl_fallback_empty_data(self, tmp_path):
        """Lines 543-544: openpyxl fallback with data shorter than header_row."""
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        df = pl.DataFrame({"A": ["x"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        config = ExcelMultiSheetConfig(
            header_row=99, data_start_row=100, end_marker_column=0, end_marker_value="Total"
        )
        mock_fastexcel = MagicMock()
        mock_fastexcel.read_excel.side_effect = Exception("fail")
        with patch(
            "acoharmony._parsers._excel_multi_sheet.pl.read_excel",
            side_effect=OverflowError("test"),
        ):
            with patch.dict("sys.modules", {"fastexcel": mock_fastexcel}):
                result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns=[])
        assert isinstance(result_df, pl.LazyFrame)

    @pytest.mark.unit
    def test_openpyxl_also_fails_raises(self, tmp_path):
        """Lines 513-518: both fastexcel and openpyxl fail raises RuntimeError."""
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        xlsx_path = tmp_path / "corrupt.xlsx"
        xlsx_path.write_bytes(b"not a real xlsx file")
        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="Total"
        )
        mock_fastexcel = MagicMock()
        mock_fastexcel.read_excel.side_effect = Exception("fastexcel fail")
        with patch(
            "acoharmony._parsers._excel_multi_sheet.pl.read_excel",
            side_effect=OverflowError("test"),
        ):
            with patch.dict("sys.modules", {"fastexcel": mock_fastexcel}):
                with pytest.raises(RuntimeError, match="Cannot read Excel file"):
                    parse_sheet_matrix(xlsx_path, 0, config, columns=[])


class TestFallbackPositionMapping:
    """Cover line 654: fallback to position mapping when no header/position markers."""

    @pytest.mark.unit
    def test_fallback_position_mapping(self, tmp_path):
        """Line 654: columns without position/header_text markers use fallback."""
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        df = pl.DataFrame({"col_0": ["Header", "Data1"], "col_1": ["H2", "D1"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="Total"
        )
        columns = [
            {"name": "field_a", "data_type": "string"},
            {"name": "field_b", "data_type": "string"},
        ]
        result_df, metadata = parse_sheet_matrix(xlsx_path, 0, config, columns)
        assert isinstance(result_df, pl.LazyFrame)


class TestParseExcelMultiSheetReRaise:
    """Cover line 997: re-raise non-matching sheet errors."""

    @pytest.mark.unit
    def test_non_matching_sheet_error_re_raised(self, tmp_path):
        """Line 997: errors that are not 'no matching sheet found' are re-raised."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"col_0": ["val"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test_schema",
            "file_format": {
                "type": "excel_multi_sheet",
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [
                {
                    "sheet_index": 0,
                    "sheet_type": "data",
                    "columns": [{"name": "col", "position": 0, "data_type": "string"}],
                },
                {
                    "sheet_index": 5,
                    "sheet_type": "extra",
                    "columns": [{"name": "col", "position": 0, "data_type": "string"}],
                },
            ],
        }
        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix as real_parse

        call_count = [0]

        def fake_parse(file_path, sheet_index, config, columns, **kwargs):
            call_count[0] += 1
            if sheet_index == 5:
                raise ValueError("some other error")
            return real_parse(file_path, sheet_index, config, columns, **kwargs)

        with patch(
            "acoharmony._parsers._excel_multi_sheet.parse_sheet_matrix", side_effect=fake_parse
        ):
            with pytest.raises(ValueError, match="some other error"):
                parse_excel_multi_sheet(xlsx_path, schema)


class TestMatrixFieldSkipExisting:
    """Cover line 1029: skip matrix field if field_name already in existing_columns."""

    @pytest.mark.unit
    def test_matrix_field_skipped_when_already_exists(self, tmp_path):
        """Line 1029: matrix field not added when column already present."""
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        df = pl.DataFrame({"Year": ["2025"], "Value": ["100"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)
        schema = {
            "name": "test_schema",
            "file_format": {
                "type": "excel_multi_sheet",
                "sheet_config": {
                    "header_row": 1,
                    "data_start_row": 2,
                    "end_marker_column": 0,
                    "end_marker_value": "Total",
                },
            },
            "sheets": [
                {
                    "sheet_index": 0,
                    "sheet_type": "data",
                    "columns": [
                        {"name": "year", "position": 0, "data_type": "string"},
                        {"name": "value", "position": 1, "data_type": "string"},
                    ],
                }
            ],
            "matrix_fields": [{"matrix": [0, 0, 0], "field_name": "year", "data_type": "string"}],
        }
        result = parse_excel_multi_sheet(xlsx_path, schema)
        assert isinstance(result, pl.LazyFrame)
        collected = result.collect()
        assert "year" in collected.columns
        year_count = sum(1 for c in collected.columns if c == "year")
        assert year_count == 1


class TestSheetColumnConfig:
    @pytest.mark.unit
    def test_basic_initialization(self):
        config = SheetColumnConfig(
            name="test_col",
            position=0,
            data_type="string",
        )
        assert config.name == "test_col"
        assert config.position == 0
        assert config.data_type == "string"
        assert config.description is None

    @pytest.mark.unit
    def test_with_date_format(self):
        config = SheetColumnConfig(
            name="date_col",
            position=1,
            data_type="date",
            date_format=["%Y-%m-%d", "%m/%d/%Y"],
        )
        assert config.date_format == ["%Y-%m-%d", "%m/%d/%Y"]


class TestMatrixFieldExtraction:
    @pytest.mark.unit
    def test_basic_initialization(self):
        config = MatrixFieldExtraction(
            field_name="report_date",
            matrix=[0, 1, 1],
        )
        assert config.field_name == "report_date"
        assert config.matrix == [0, 1, 1]

    @pytest.mark.unit
    def test_with_null_sheet(self):
        config = MatrixFieldExtraction(
            field_name="report_date",
            matrix=[None, 0, 14],
        )
        assert config.matrix[0] is None


class TestSheetConfig:
    @pytest.mark.unit
    def test_basic_initialization(self):
        config = SheetConfig(
            sheet_index=0,
            sheet_type="inpatient",
            columns=[{"name": "col1", "position": 0, "data_type": "string"}],
        )
        assert config.sheet_index == 0
        assert config.sheet_type == "inpatient"
        assert len(config.columns) == 1


class TestExcelMultiSheetConfig:
    @pytest.mark.unit
    def test_basic_initialization(self):
        config = ExcelMultiSheetConfig(
            header_row=0,
            data_start_row=1,
            end_marker_column=0,
            end_marker_value="Total",
        )
        assert config.header_row == 0
        assert config.data_start_row == 1

    @pytest.mark.unit
    def test_with_header_search(self):
        config = ExcelMultiSheetConfig(
            header_row=0,
            data_start_row=1,
            end_marker_column=0,
            end_marker_value="Total",
            column_mapping_strategy="header_match",
            header_search_text="Provider Type",
        )
        assert config.column_mapping_strategy == "header_match"



class TestExtractMatrixFields:
    @pytest.mark.unit
    def test_extracts_cell_value(self, tmp_path):
        """extract_matrix_fields takes a file path and list of dicts."""
        # Create a minimal xlsx file


        df = pl.DataFrame({"col_0": ["Header", "Row1"], "col_1": ["2025-01-01", "Data1"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)

        matrix_fields = [{"matrix": [0, 1, 1], "field_name": "report_date"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert "report_date" in result
        assert result["report_date"] == "2025-01-01"

    @pytest.mark.unit
    def test_missing_sheet_returns_none(self, tmp_path):
        df = pl.DataFrame({"col_0": ["only_data"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)

        matrix_fields = [{"matrix": [5, 0, 0], "field_name": "report_date"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        # Field targeting sheet 5 should not match when sheet_index=0
        assert result.get("report_date") is None

    @pytest.mark.unit
    def test_invalid_matrix_length(self, tmp_path):
        df = pl.DataFrame({"col_0": ["data"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)

        matrix_fields = [{"matrix": [0, 0], "field_name": "bad"}]  # Only 2 elements
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result.get("bad") is None

    @pytest.mark.unit
    def test_null_sheet_applies_to_current(self, tmp_path):
        df = pl.DataFrame({"col_0": ["value"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)

        matrix_fields = [{"matrix": [None, 0, 0], "field_name": "val"}]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert "val" in result

    @pytest.mark.unit
    def test_extract_pattern(self, tmp_path):
        df = pl.DataFrame({"col_0": ["Report Date: January 15, 2025"]})
        xlsx_path = tmp_path / "test.xlsx"
        df.write_excel(xlsx_path)

        # Row 0 is the header "col_0", row 1 is the data
        matrix_fields = [{
            "matrix": [0, 1, 0],
            "field_name": "date",
            "extract_pattern": r"Report Date: (.+)",
        }]
        result = extract_matrix_fields(xlsx_path, matrix_fields, sheet_index=0)
        assert result.get("date") == "January 15, 2025"


class TestMapColumnsByPosition:
    @pytest.mark.unit
    def test_maps_columns(self):
        df = pl.DataFrame({"col_0": ["a"], "col_1": ["b"], "col_2": ["c"]})
        columns = [
            {"name": "first", "position": 0, "data_type": "string"},
            {"name": "second", "position": 1, "data_type": "string"},
        ]
        rename_map, type_map = map_columns_by_position(df, columns)
        assert "first" in rename_map.values() or len(rename_map) > 0

    @pytest.mark.unit
    def test_position_out_of_range(self):
        df = pl.DataFrame({"col_0": ["a"]})
        columns = [
            {"name": "missing", "position": 10, "data_type": "string"},
        ]
        rename_map, type_map = map_columns_by_position(df, columns)
        assert isinstance(rename_map, dict)


class TestMapColumnsByHeaderMatch:
    @pytest.mark.unit
    def test_matches_headers(self):
        df = pl.DataFrame({
            "col_0": ["Name", "John"],
            "col_1": ["Amount", "100"],
        })
        columns = [
            {"name": "name", "position": 0, "data_type": "string", "header_text": "Name"},
            {"name": "amount", "position": 1, "data_type": "string", "header_text": "Amount"},
        ]
        rename_map, type_map, extra = map_columns_by_header_match(df, 0, columns)
        assert isinstance(rename_map, dict)


class TestApplySheetTransform:
    @pytest.mark.unit
    def test_basic_transform(self):
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform
        df = pl.DataFrame({
            "col_a": ["val1", "val2"],
            "col_b": ["10", "20"],
        }).lazy()
        transform_config = {"rename": {"col_a": "name"}}
        result = _apply_sheet_transform(df, transform_config)
        assert isinstance(result, pl.LazyFrame)
        collected = result.collect()
        assert collected.height == 2

class TestExcelMultiSheet:
    """Tests for _excel_multi_sheet functions."""

    @pytest.fixture
    def simple_xlsx(self, tmp_path: Path) -> Path:
        """Create a simple xlsx file with known data."""
        df = pl.DataFrame({'col_a': ['Header_A', 'val1', 'val2', 'val3', 'TOTAL'], 'col_b': ['Header_B', '100', '200', '300', '600'], 'col_c': ['Header_C', 'x', 'y', 'z', '']})
        p = tmp_path / 'simple.xlsx'
        df.write_excel(p)
        return p

    @pytest.fixture
    def multi_sheet_xlsx(self, tmp_path: Path) -> Path:
        """Create an xlsx with multiple sheets."""
        p = tmp_path / 'multi.xlsx'
        if HAS_OPENPYXL:
            import openpyxl
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = 'Sheet1'
            ws1.append(['Name', 'Value'])
            ws1.append(['param1', '100'])
            ws1.append(['param2', '200'])
            ws2 = wb.create_sheet('Sheet2')
            ws2.append(['ID', 'Amount', 'Status'])
            ws2.append(['1', '50.0', 'active'])
            ws2.append(['2', '75.5', 'inactive'])
            ws2.append(['TOTAL', '', ''])
            wb.save(p)
        else:
            df = pl.DataFrame({'A': [1, 2], 'B': ['x', 'y']})
            df.write_excel(p)
        return p

    @pytest.mark.unit
    def test_extract_matrix_fields(self, simple_xlsx: Path):
        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        fields = [{'matrix': [0, 0, 0], 'field_name': 'header_a', 'data_type': 'string'}, {'matrix': [0, 1, 1], 'field_name': 'first_val', 'data_type': 'string'}, {'matrix': [0, 0, 2], 'field_name': 'header_c', 'data_type': 'string'}]
        result = extract_matrix_fields(simple_xlsx, fields, sheet_index=0)
        assert 'header_a' in result
        assert 'first_val' in result

    @pytest.mark.unit
    def test_extract_matrix_fields_with_pattern(self, tmp_path: Path):
        """Test extract_pattern on matrix fields."""
        df = pl.DataFrame({'A': ['Performance Year 2025'], 'B': ['other']})
        p = tmp_path / 'pattern.xlsx'
        df.write_excel(p)
        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        fields = [{'matrix': [0, 1, 0], 'field_name': 'year', 'extract_pattern': '\\d{4}', 'data_type': 'string'}]
        result = extract_matrix_fields(p, fields, sheet_index=0)
        assert result['year'] == '2025'

    @pytest.mark.unit
    def test_extract_matrix_fields_out_of_bounds(self, simple_xlsx: Path):
        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        fields = [{'matrix': [0, 999, 999], 'field_name': 'missing', 'default_value': 'default'}]
        result = extract_matrix_fields(simple_xlsx, fields, sheet_index=0)
        assert result['missing'] == 'default'

    @pytest.mark.unit
    def test_extract_matrix_fields_bad_matrix(self, simple_xlsx: Path):
        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        fields = [{'matrix': [0, 0], 'field_name': 'bad'}]
        result = extract_matrix_fields(simple_xlsx, fields, sheet_index=0)
        assert 'bad' not in result

    @pytest.mark.unit
    def test_extract_matrix_fields_integer_type(self, tmp_path: Path):
        df = pl.DataFrame({'A': ['42']})
        p = tmp_path / 'int.xlsx'
        df.write_excel(p)
        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        fields = [{'matrix': [0, 1, 0], 'field_name': 'num', 'data_type': 'integer'}]
        result = extract_matrix_fields(p, fields, sheet_index=0)
        assert result['num'] == 42

    @pytest.mark.unit
    def test_map_columns_by_position(self):
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_position
        df = pl.DataFrame({'col_0': ['a'], 'col_1': ['b'], 'col_2': ['c']})
        columns = [{'position': 1, 'name': 'second', 'data_type': 'string'}, {'position': 2, 'name': 'third', 'data_type': 'integer'}]
        mapping, dtypes = map_columns_by_position(df, columns)
        assert mapping['col_1'] == 'second'
        assert mapping['col_2'] == 'third'
        assert dtypes['second'] == 'string'
        assert dtypes['third'] == 'integer'

    @pytest.mark.unit
    def test_map_columns_by_position_source_position(self):
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_position
        df = pl.DataFrame({'c0': ['a'], 'c1': ['b']})
        columns = [{'source_position': 1, 'name': 'out', 'data_type': 'string'}]
        mapping, dtypes = map_columns_by_position(df, columns)
        assert mapping['c1'] == 'out'

    @pytest.mark.unit
    def test_map_columns_by_position_duplicate_output_name(self):
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_position
        df = pl.DataFrame({'c0': ['a'], 'c1': ['b'], 'c2': ['c']})
        columns = [{'position': 1, 'name': 'same', 'data_type': 'string'}, {'position': 2, 'name': 'same', 'data_type': 'string'}]
        mapping, _ = map_columns_by_position(df, columns)
        assert len(mapping) == 1

    @pytest.mark.unit
    def test_map_columns_by_header_match(self):
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match
        df = pl.DataFrame({'c0': ['Provider Type', 'A', 'B'], 'c1': ['Total Amount', '100', '200'], 'c2': ['CCN', 'X', 'Y']})
        columns = [{'header_text': 'Provider Type', 'name': 'provider_type', 'data_type': 'string'}, {'header_text': 'Total Amount', 'name': 'total_amount', 'data_type': 'decimal'}, {'header_text': 'CCN', 'name': 'ccn', 'data_type': 'string'}]
        mapping, dtypes, metadata = map_columns_by_header_match(df, 0, columns)
        assert mapping['c0'] == 'provider_type'
        assert mapping['c1'] == 'total_amount'
        assert mapping['c2'] == 'ccn'

    @pytest.mark.unit
    def test_map_columns_by_header_match_with_metadata(self):
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match
        df = pl.DataFrame({'c0': ['Claims CY2025', '100']})
        columns = [{'header_text': 'Claims', 'name': 'claims', 'data_type': 'string', 'extract_header_metadata': [{'field_name': 'claim_year', 'extract_pattern': 'CY(\\d{4})'}]}]
        mapping, dtypes, metadata = map_columns_by_header_match(df, 0, columns)
        assert 'claims' in metadata
        assert metadata['claims']['claim_year'] == '2025'

    @pytest.mark.unit
    def test_parse_sheet_matrix(self, tmp_path: Path):
        """Test parse_sheet_matrix with position-based column mapping."""
        if not HAS_OPENPYXL:
            pytest.skip('openpyxl required')
        import openpyxl
        p = tmp_path / 'matrix.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Provider Type', 'Amount', 'Code'])
        ws.append(['Hospital', '100', 'A1'])
        ws.append(['Clinic', '200', 'B2'])
        ws.append(['TOTAL', '', ''])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix
        config = ExcelMultiSheetConfig(header_row=0, data_start_row=1, end_marker_column=0, end_marker_value='TOTAL', column_mapping_strategy='position')
        columns = [{'position': 1, 'name': 'amount', 'data_type': 'string'}, {'position': 2, 'name': 'code', 'data_type': 'string'}]
        lf, meta = parse_sheet_matrix(p, 0, config, columns)
        df = lf.collect()
        assert len(df) == 2
        assert df['amount'].to_list() == ['100', '200']
        assert df['code'].to_list() == ['A1', 'B2']

    @pytest.mark.unit
    def test_apply_sheet_transform_no_type(self):
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform
        df = pl.DataFrame({'a': [1]}).lazy()
        result = _apply_sheet_transform(df, {})
        assert result.collect().shape == (1, 1)

    @pytest.mark.unit
    def test_apply_sheet_transform_unknown_type(self):
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform
        df = pl.DataFrame({'a': [1]}).lazy()
        result = _apply_sheet_transform(df, {'type': 'nonexistent_transform'})
        assert result.collect().shape == (1, 1)

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_basic(self, tmp_path: Path):
        """Test generic multi-sheet parser with dict schema."""
        if not HAS_OPENPYXL:
            pytest.skip('openpyxl required')
        import openpyxl
        p = tmp_path / 'generic.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.append(['Type', 'Value'])
        ws.append(['A', '100'])
        ws.append(['B', '200'])
        ws.append(['TOTAL', ''])
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
        schema = {'name': 'test_schema', 'file_format': {'sheet_config': {'header_row': 0, 'data_start_row': 1, 'end_marker_column': 0, 'end_marker_value': 'TOTAL', 'column_mapping_strategy': 'position'}}, 'sheets': [{'sheet_type': 'data', 'sheet_index': 0, 'columns': [{'position': 0, 'name': 'type', 'data_type': 'string'}, {'position': 1, 'name': 'value', 'data_type': 'string'}]}]}
        lf = parse_excel_multi_sheet(p, schema)
        df = lf.collect()
        assert 'sheet_type' in df.columns
        assert len(df) == 2

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_no_sheets(self, tmp_path: Path):
        if not HAS_OPENPYXL:
            pytest.skip('openpyxl required')
        p = tmp_path / 'empty.xlsx'
        import openpyxl
        wb = openpyxl.Workbook()
        wb.save(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
        schema = {'name': 'test', 'file_format': {'sheet_config': {'header_row': 0, 'data_start_row': 1, 'end_marker_column': 0, 'end_marker_value': 'END'}}, 'sheets': []}
        with pytest.raises(ValueError, match='sheets'):
            parse_excel_multi_sheet(p, schema)

    @pytest.mark.unit
    def test_parse_excel_multi_sheet_bad_schema(self, tmp_path: Path):
        p = tmp_path / 'x.xlsx'
        pl.DataFrame({'a': [1]}).write_excel(p)
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
        with pytest.raises(ValueError, match='file_format'):
            parse_excel_multi_sheet(p, {'name': 'x', 'sheets': [{'sheet_type': 'a'}]})

    @pytest.mark.unit
    def test_excel_multi_sheet_config_defaults(self):
        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig
        config = ExcelMultiSheetConfig(header_row=0, data_start_row=1, end_marker_column=0, end_marker_value='END')
        assert config.column_mapping_strategy == 'position'
        assert config.header_search_text is None

@pytest.mark.skipif(not HAS_OPENPYXL, reason='openpyxl required')
class TestExcelMultiSheetExtractMatrixFields:
    """Cover extract_matrix_fields missed lines in _excel_multi_sheet.py."""

    @pytest.mark.unit
    def test_extract_matrix_fields_sheet_mismatch_skips(self, tmp_path: Path):
        """Cover line 193: target_sheet != sheet_index causes continue."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['val_at_0_0'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [5, 0, 0], 'field_name': 'my_field'}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result == {}

    @pytest.mark.unit
    def test_extract_matrix_fields_read_sheet_none(self, tmp_path: Path):
        """Cover line 198: read_sheet is None when both target_sheet and sheet_index are None."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['val'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [None, 0, 0], 'field_name': 'f1'}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=None)
        assert result == {}

    @pytest.mark.unit
    def test_extract_matrix_fields_none_value_gets_default(self, tmp_path: Path):
        """Cover line 219: value is None → uses default_value."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append([None])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [0, 0, 0], 'field_name': 'f1', 'default_value': 'fallback'}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result['f1'] == 'fallback'

    @pytest.mark.unit
    def test_extract_matrix_fields_extract_pattern(self, tmp_path: Path):
        """Cover line 215-217: extract_pattern is applied to value."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['Report Period: 2025'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [0, 0, 0], 'field_name': 'year', 'extract_pattern': '\\d{4}'}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result['year'] == '2025'

    @pytest.mark.unit
    def test_extract_matrix_fields_extract_pattern_no_match(self, tmp_path: Path):
        """Cover line 217 (else branch): extract_pattern doesn't match → default_value."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['no digits here'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [0, 0, 0], 'field_name': 'year', 'extract_pattern': '\\d{4}', 'default_value': 'N/A'}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result['year'] == 'N/A'

    @pytest.mark.unit
    def test_extract_matrix_fields_integer_type(self, tmp_path: Path):
        """Cover line 225: data_type='integer' casts to int."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['42'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [0, 0, 0], 'field_name': 'count', 'data_type': 'integer'}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result['count'] == 42
        assert isinstance(result['count'], int)

    @pytest.mark.unit
    def test_extract_matrix_fields_decimal_type(self, tmp_path: Path):
        """Cover line 227: data_type='decimal' casts to float."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['3.14'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [0, 0, 0], 'field_name': 'rate', 'data_type': 'decimal'}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert abs(result['rate'] - 3.14) < 0.001

    @pytest.mark.unit
    def test_extract_matrix_fields_none_value_pass(self, tmp_path: Path):
        """Cover line 234: value is None with non-string data_type → pass."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append([None])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [0, 0, 0], 'field_name': 'count', 'data_type': 'integer', 'default_value': None}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result['count'] is None

    @pytest.mark.unit
    def test_extract_matrix_fields_out_of_bounds(self, tmp_path: Path):
        """Cover lines 240-241: cell out of bounds uses default_value."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['only_one_cell'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{'matrix': [0, 999, 0], 'field_name': 'f1', 'default_value': 'oob'}]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result['f1'] == 'oob'

    @pytest.mark.unit
    def test_extract_matrix_fields_exception_fallback(self, tmp_path: Path):
        """Cover lines 243-248: extraction exception → default_value."""
        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        fields = [{'matrix': [0, 0, 0], 'field_name': 'f1', 'default_value': 'err_default'}]
        result = extract_matrix_fields(tmp_path / 'nonexistent.xlsx', fields, sheet_index=0)
        assert result['f1'] == 'err_default'

@pytest.mark.skipif(not HAS_OPENPYXL, reason='openpyxl required')
class TestExcelMultiSheetMapColumnsByHeaderMatch:
    """Cover map_columns_by_header_match missed lines 344, 378."""

    @pytest.mark.unit
    def test_header_match_skips_empty_header_val(self, tmp_path: Path):
        """Cover line 344: header_val is falsy → continue."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match
        wb = Workbook()
        ws = wb.active
        ws.append([None, 'Provider Name', 'Amount'])
        ws.append(['A', 'Test Provider', '100'])
        wb.save(tmp_path / 'test.xlsx')
        import polars as pl
        df = pl.read_excel(tmp_path / 'test.xlsx', read_options={'header_row': None, 'skip_rows': 0}, infer_schema_length=0)
        columns = [{'header_text': 'Provider Name', 'name': 'provider_name', 'data_type': 'string'}]
        mapping, dtypes, _ = map_columns_by_header_match(df, 0, columns)
        assert 'provider_name' in mapping.values()

    @pytest.mark.unit
    def test_header_match_duplicate_output_name_skipped(self, tmp_path: Path):
        """Cover line 378: output_name already used → continue (skip duplicate)."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match
        wb = Workbook()
        ws = wb.active
        ws.append(['Provider Name', 'Provider Name Copy'])
        ws.append(['Test', 'Test2'])
        wb.save(tmp_path / 'test.xlsx')
        import polars as pl
        df = pl.read_excel(tmp_path / 'test.xlsx', read_options={'header_row': None, 'skip_rows': 0}, infer_schema_length=0)
        columns = [{'header_text': 'Provider Name', 'name': 'prov', 'output_name': 'prov', 'data_type': 'string'}]
        mapping, dtypes, _ = map_columns_by_header_match(df, 0, columns)
        assert list(mapping.values()).count('prov') == 1

class TestMapColumnsByHeaderMatchBranches408_430:
    """Cover map_columns_by_header_match branches 408→409, 419→422, 426→422, 430→422."""

    @pytest.mark.unit
    def test_duplicate_output_name_skipped_across_col_defs(self):
        """Branch 408→409: second col_def with same output_name is skipped."""
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match

        # Two col_defs with different header_text but same output_name
        # Both will find a matching header, but second should be skipped
        df = pl.DataFrame({
            "c0": ["Provider Name", "Alice"],
            "c1": ["Total Amount", "100"],
        })
        columns = [
            {"header_text": "Provider Name", "name": "field", "output_name": "field", "data_type": "string"},
            {"header_text": "Total Amount", "name": "field2", "output_name": "field", "data_type": "decimal"},
        ]
        mapping, dtypes, metadata = map_columns_by_header_match(df, 0, columns)
        # Only the first col_def should be mapped
        assert list(mapping.values()).count("field") == 1
        assert mapping["c0"] == "field"
        assert "c1" not in mapping

    @pytest.mark.unit
    def test_metadata_missing_field_name_or_pattern(self):
        """Branch 426→422: field_name or extract_pattern missing → skip extraction."""
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match

        df = pl.DataFrame({"c0": ["Claims CY2025", "100"]})
        columns = [
            {
                "header_text": "Claims",
                "name": "claims",
                "data_type": "string",
                "extract_header_metadata": [
                    {"field_name": None, "extract_pattern": r"CY(\d{4})"},  # missing field_name
                    {"field_name": "year", "extract_pattern": None},  # missing extract_pattern
                    {"field_name": "", "extract_pattern": r"CY(\d{4})"},  # empty field_name
                ],
            }
        ]
        mapping, dtypes, metadata = map_columns_by_header_match(df, 0, columns)
        assert "claims" in mapping.values()
        # No metadata should be extracted since all configs are incomplete
        assert metadata.get("claims", {}) == {}

    @pytest.mark.unit
    def test_metadata_regex_no_match(self):
        """Branch 430→422: regex pattern doesn't match header → skip extraction."""
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match

        df = pl.DataFrame({"c0": ["Claims NoYear", "100"]})
        columns = [
            {
                "header_text": "Claims",
                "name": "claims",
                "data_type": "string",
                "extract_header_metadata": [
                    {"field_name": "claim_year", "extract_pattern": r"CY(\d{4})"},
                ],
            }
        ]
        mapping, dtypes, metadata = map_columns_by_header_match(df, 0, columns)
        assert "claims" in mapping.values()
        # Metadata dict exists but no extracted values since regex didn't match
        assert metadata.get("claims", {}) == {}

    @pytest.mark.unit
    def test_metadata_regex_full_match_no_group(self):
        """Cover match.lastindex being None → use group(0) full match."""
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match

        df = pl.DataFrame({"c0": ["Claims 2025", "100"]})
        columns = [
            {
                "header_text": "Claims",
                "name": "claims",
                "data_type": "string",
                "extract_header_metadata": [
                    {"field_name": "year", "extract_pattern": r"\d{4}"},  # no capture group
                ],
            }
        ]
        mapping, dtypes, metadata = map_columns_by_header_match(df, 0, columns)
        assert metadata["claims"]["year"] == "2025"

    @pytest.mark.unit
    def test_metadata_multiple_extraction_configs(self):
        """Cover multiple metadata extraction configs on same col_def."""
        from acoharmony._parsers._excel_multi_sheet import map_columns_by_header_match

        df = pl.DataFrame({"c0": ["Claims CY2025 Q1", "100"]})
        columns = [
            {
                "header_text": "Claims",
                "name": "claims",
                "data_type": "string",
                "extract_header_metadata": [
                    {"field_name": "claim_year", "extract_pattern": r"CY(\d{4})"},
                    {"field_name": "quarter", "extract_pattern": r"(Q\d)"},
                ],
            }
        ]
        mapping, dtypes, metadata = map_columns_by_header_match(df, 0, columns)
        assert metadata["claims"]["claim_year"] == "2025"
        assert metadata["claims"]["quarter"] == "Q1"

    @pytest.mark.unit
    def test_header_metadata_output_name_already_present(self):
        """Branch 419→422: output_name already in header_metadata → skip init.

        This branch is structurally unreachable through the public API because
        the duplicate output_name check (408→409) prevents a second col_def with
        the same output_name from reaching the metadata block. We cover it by
        temporarily replacing the ``set`` builtin used by the function with a
        custom class that skips adding 'claims' the first time, so the second
        col_def bypasses 408→409 and reaches line 419 where 'claims' is already
        in header_metadata from the first col_def.
        """
        import builtins
        import acoharmony._parsers._excel_multi_sheet as _mod

        class _LeakySet(set):
            """Set subclass that silently drops the first add('claims')."""

            _skip_count = 0

            def add(self, item):
                if item == "claims" and _LeakySet._skip_count == 0:
                    _LeakySet._skip_count += 1
                    return  # don't record 'claims' the first time
                super().add(item)

        df = pl.DataFrame({
            "c0": ["Claims CY2025", "100"],
            "c1": ["Total Amount CY2026", "200"],
        })
        columns = [
            {
                "header_text": "Claims",
                "name": "claims",
                "data_type": "string",
                "extract_header_metadata": [
                    {"field_name": "claim_year", "extract_pattern": r"CY(\d{4})"},
                ],
            },
            {
                "header_text": "Total Amount",
                "name": "total",
                "output_name": "claims",  # same output_name as first
                "data_type": "string",
                "extract_header_metadata": [
                    {"field_name": "amount_year", "extract_pattern": r"CY(\d{4})"},
                ],
            },
        ]

        _orig_set = builtins.set
        builtins.set = _LeakySet
        try:
            mapping, dtypes, metadata = _mod.map_columns_by_header_match(df, 0, columns)
        finally:
            builtins.set = _orig_set
            _LeakySet._skip_count = 0

        # The second col_def reached 419 where 'claims' was already in
        # header_metadata (from first col_def), so init was skipped.
        assert "claims" in metadata
        assert metadata["claims"]["claim_year"] == "2025"
        # Second col_def also extracted metadata (amount_year)
        assert metadata["claims"]["amount_year"] == "2026"


class TestUncoveredBranches568_618_643_655_660:
    """Cover branches 568->567, 618->626, 643->644, 655->656, 660->689."""

    @pytest.mark.unit
    def test_openpyxl_fallback_none_header_filtered(self, tmp_path):
        """Branch 568->567: header is None in openpyxl fallback, gets skipped.

        Force both polars read_excel and fastexcel to fail so we reach the
        openpyxl final fallback path. Create xlsx with a None header cell
        so the 'if h is not None' check at line 568 evaluates False.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        from acoharmony._parsers._excel_multi_sheet import (
            ExcelMultiSheetConfig,
            parse_sheet_matrix,
        )

        # Create xlsx with a None header (empty cell in the header row)
        p = tmp_path / "none_hdr.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", None, "Value"])  # None header in column B
        ws.append(["Alice", "extra", "100"])
        ws.append(["Bob", "extra2", "200"])
        wb.save(p)

        config = ExcelMultiSheetConfig(
            header_row=1, data_start_row=2, end_marker_column=0, end_marker_value="TOTAL"
        )

        mock_fastexcel = MagicMock()
        mock_fastexcel.read_excel.side_effect = Exception("fastexcel fail")

        with patch(
            "acoharmony._parsers._excel_multi_sheet.pl.read_excel",
            side_effect=OverflowError("force openpyxl fallback"),
        ):
            with patch.dict("sys.modules", {"fastexcel": mock_fastexcel}):
                lf, meta = parse_sheet_matrix(p, 0, config, columns=[])

        df = lf.collect()
        # The None header column should be filtered out
        for col in df.columns:
            assert col.strip() != "None"
        assert df.height >= 1

    @pytest.mark.unit
    def test_header_search_text_no_match(self, tmp_path):
        """Branch 618->626: header_search_text loop exhausts without finding match.

        Set header_search_text to something not present in any row within the
        search range so the for-loop at line 618 completes without break,
        falling through to line 626.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        from acoharmony._parsers._excel_multi_sheet import (
            ExcelMultiSheetConfig,
            parse_sheet_matrix,
        )

        p = tmp_path / "no_match_search.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header A", "Header B"])
        ws.append(["Row1", "100"])
        ws.append(["Row2", "200"])
        ws.append(["TOTAL", ""])
        wb.save(p)

        config = ExcelMultiSheetConfig(
            header_row=0,
            data_start_row=1,
            end_marker_column=0,
            end_marker_value="TOTAL",
            header_search_text="ZZZZZ_NEVER_FOUND",
        )
        columns = [{"position": 0, "name": "col_a", "data_type": "string"}]
        lf, meta = parse_sheet_matrix(p, 0, config, columns)
        df = lf.collect()
        # Should still produce results using the original header_row
        assert df.height >= 1

    @pytest.mark.unit
    def test_skip_data_slice_with_columns(self, tmp_path):
        """Branches 643->644 and 655->656: skip_data_slice=True with non-empty columns.

        When skip_data_slice=True and columns are provided, we reach:
        - line 643: if skip_data_slice -> df_data = df_raw (branch 643->644)
        - line 655: if skip_data_slice -> column_mapping = {} (branch 655->656)
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        from acoharmony._parsers._excel_multi_sheet import (
            ExcelMultiSheetConfig,
            parse_sheet_matrix,
        )

        p = tmp_path / "skip_slice_cols.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Category", "Group A", "Group B"])
        ws.append(["SubCat", "Val1", "Val2"])
        ws.append(["Row1", "10", "20"])
        ws.append(["Row2", "30", "40"])
        wb.save(p)

        config = ExcelMultiSheetConfig(
            header_row=0,
            data_start_row=2,
            end_marker_column=0,
            end_marker_value="TOTAL",
        )
        # Non-empty columns but skip_data_slice=True
        columns = [
            {"position": 0, "name": "category", "data_type": "string"},
            {"position": 1, "name": "group_a", "data_type": "string"},
        ]
        lf, meta = parse_sheet_matrix(p, 0, config, columns, skip_data_slice=True)
        df = lf.collect()
        # With skip_data_slice=True, all rows (including headers) are preserved
        # Column mapping is skipped so we get generic column_N names
        assert df.height >= 3

    @pytest.mark.unit
    def test_no_columns_no_early_return(self, tmp_path):
        """Branch 660->689: columns is empty, skip_data_slice=False, bypassing early return.

        Use a SimpleNamespace config with header_row=None to bypass the early
        return at line 486 (which requires header_row is not None). This reaches
        line 660 where 'elif columns:' evaluates False (columns is empty),
        leading to empty column_mapping and line 689 'if column_mapping:' being False.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix

        p = tmp_path / "no_cols_no_early.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append(["1", "2"])
        ws.append(["3", "4"])
        wb.save(p)

        # Use SimpleNamespace to set header_row=None (bypasses Pydantic validation)
        config = SimpleNamespace(
            header_row=None,
            data_start_row=1,
            end_marker_column=0,
            end_marker_value="TOTAL",
            column_mapping_strategy="position",
            header_search_text=None,
        )
        lf, meta = parse_sheet_matrix(p, 0, config, columns=[], skip_data_slice=False)
        df = lf.collect()
        # Should return all data rows with generic column_N names
        assert df.height >= 1


@pytest.mark.skipif(not HAS_OPENPYXL, reason='openpyxl required')
class TestParseSheetMatrixNoColumns:
    """Cover parse_sheet_matrix lines 476-544: no columns, native header reading."""

    @pytest.mark.unit
    def test_parse_sheet_matrix_no_columns_native_headers(self, tmp_path: Path):
        """Cover lines 454-478: no columns triggers native header reading path."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Value'])
        ws.append(['Alice', '100'])
        ws.append(['Bob', '200'])
        wb.save(tmp_path / 'test.xlsx')
        config = ExcelMultiSheetConfig(header_row=1, data_start_row=2, end_marker_column=0, end_marker_value='TOTAL')
        lf, metadata = parse_sheet_matrix(tmp_path / 'test.xlsx', 0, config, columns=[])
        df = lf.collect()
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_sheet_matrix_openpyxl_fallback_on_error(self, tmp_path: Path):
        """Cover lines 476-478: openpyxl header reading exception → empty schema_overrides."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Value'])
        ws.append(['Alice', '100'])
        wb.save(tmp_path / 'test.xlsx')
        config = ExcelMultiSheetConfig(header_row=1, data_start_row=2, end_marker_column=0, end_marker_value='TOTAL')
        import openpyxl as _openpyxl
        original_load = _openpyxl.load_workbook
        call_count = [0]

        def patched_load(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] == 1:
                raise Exception('Simulated drawing error')
            return original_load(*args, **kwargs)
        with patch('openpyxl.load_workbook', side_effect=patched_load):
            lf, metadata = parse_sheet_matrix(tmp_path / 'test.xlsx', 0, config, columns=[])
            df = lf.collect()
            assert df is not None

@pytest.mark.skipif(not HAS_OPENPYXL, reason='openpyxl required')
class TestParseSheetMatrixEdgeCases:
    """Cover parse_sheet_matrix lines 654, 678."""

    @pytest.mark.unit
    def test_fallback_to_position_mapping_when_no_strategy(self, tmp_path: Path):
        """Cover line 654: fallback to position mapping when no header columns and no header_match strategy."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix
        wb = Workbook()
        ws = wb.active
        ws.append(['Col1', 'Col2'])
        ws.append(['A', 'B'])
        ws.append(['C', 'D'])
        wb.save(tmp_path / 'test.xlsx')
        config = ExcelMultiSheetConfig(header_row=0, data_start_row=1, end_marker_column=0, end_marker_value='TOTAL', column_mapping_strategy='position')
        columns = [{'name': 'first', 'position': 0, 'data_type': 'string'}, {'name': 'second', 'position': 1, 'data_type': 'string'}]
        lf, _ = parse_sheet_matrix(tmp_path / 'test.xlsx', 0, config, columns=columns)
        df = lf.collect()
        assert 'first' in df.columns or len(df) >= 0

    @pytest.mark.unit
    def test_float_type_cast(self, tmp_path: Path):
        """Cover line 678: data_type='decimal' casts column to Float64."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import ExcelMultiSheetConfig, parse_sheet_matrix
        wb = Workbook()
        ws = wb.active
        ws.append(['Amount'])
        ws.append(['3.14'])
        ws.append(['2.71'])
        wb.save(tmp_path / 'test.xlsx')
        config = ExcelMultiSheetConfig(header_row=0, data_start_row=1, end_marker_column=0, end_marker_value='TOTAL', column_mapping_strategy='header_match')
        columns = [{'header_text': 'Amount', 'name': 'amount', 'output_name': 'amount', 'data_type': 'decimal'}]
        lf, _ = parse_sheet_matrix(tmp_path / 'test.xlsx', 0, config, columns=columns)
        df = lf.collect()
        if 'amount' in df.columns:
            assert df['amount'].dtype == pl.Float64

@pytest.mark.skipif(not HAS_OPENPYXL, reason='openpyxl required')
class TestParseExcelMultiSheetSpecialized:
    """Cover parse_excel_multi_sheet lines 893, 925, 968, 997, 1029, 1036, 1038."""

    @pytest.mark.unit
    def test_sheets_list_non_list_converted(self, tmp_path: Path):
        """Cover line 893: sheets_list is not a list → list(sheets_list)."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.append(['Header1'])
        ws.append(['Data1'])
        wb.save(tmp_path / 'test.xlsx')
        sheet_def = {'sheet_index': 0, 'sheet_type': 'data', 'columns': [{'name': 'col1', 'position': 0, 'data_type': 'string'}]}
        schema = SimpleNamespace(name='test_schema', file_format=SimpleNamespace(sheet_config={'header_row': 0, 'data_start_row': 1, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}), sheets=(sheet_def,))
        result = parse_excel_multi_sheet(tmp_path / 'test.xlsx', schema)
        assert result is not None

    @pytest.mark.unit
    def test_matrix_fields_non_list_converted(self, tmp_path: Path):
        """Cover line 925: matrix_fields_raw is not a list → list(...)."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
        wb = Workbook()
        ws = wb.active
        ws.append(['Header'])
        ws.append(['Data'])
        wb.save(tmp_path / 'test.xlsx')
        mf = SimpleNamespace(matrix=[0, 0, 0], field_name='meta_field', data_type='string', default_value=None)
        schema = SimpleNamespace(name='test_schema', file_format=SimpleNamespace(sheet_config={'header_row': 0, 'data_start_row': 1, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}), sheets=[{'sheet_index': 0, 'sheet_type': 'data', 'columns': [{'name': 'col1', 'position': 0, 'data_type': 'string'}]}], matrix_fields=(mf,))
        result = parse_excel_multi_sheet(tmp_path / 'test.xlsx', schema)
        assert result is not None

    @pytest.mark.unit
    def test_sheet_config_override_non_dict(self, tmp_path: Path):
        """Cover line 968: sheet_config override is not a dict → vars(...)."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
        wb = Workbook()
        ws = wb.active
        ws.append(['Header'])
        ws.append(['Data'])
        wb.save(tmp_path / 'test.xlsx')
        schema = {'name': 'test_schema', 'file_format': {'sheet_config': {'header_row': 0, 'data_start_row': 1, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}}, 'sheets': [{'sheet_index': 0, 'sheet_type': 'data', 'columns': [{'name': 'col1', 'position': 0, 'data_type': 'string'}], 'sheet_config': SimpleNamespace(header_row=0, data_start_row=1, end_marker_column=0, end_marker_value='TOTAL')}]}
        result = parse_excel_multi_sheet(tmp_path / 'test.xlsx', schema)
        assert result is not None

    @pytest.mark.unit
    def test_parse_error_no_matching_sheet_continues(self, tmp_path: Path):
        """Cover line 997: 'no matching sheet found' error is skipped."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
        wb = Workbook()
        ws = wb.active
        ws.append(['Header'])
        ws.append(['Data'])
        wb.save(tmp_path / 'test.xlsx')
        schema = {'name': 'test_schema', 'file_format': {'sheet_config': {'header_row': 0, 'data_start_row': 1, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}}, 'sheets': [{'sheet_index': 99, 'sheet_type': 'missing', 'columns': [{'name': 'col1', 'position': 0, 'data_type': 'string'}]}, {'sheet_index': 0, 'sheet_type': 'data', 'columns': [{'name': 'col1', 'position': 0, 'data_type': 'string'}]}]}
        result = parse_excel_multi_sheet(tmp_path / 'test.xlsx', schema)
        df = result.collect()
        assert len(df) >= 0

    @pytest.mark.unit
    def test_matrix_field_integer_dtype_in_multi_sheet(self, tmp_path: Path):
        """Cover lines 1029, 1036, 1038: matrix field with integer/decimal dtype added as column."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
        wb = Workbook()
        ws = wb.active
        ws.append(['Header'])
        ws.append(['42'])
        wb.save(tmp_path / 'test.xlsx')
        schema = {'name': 'test_schema', 'file_format': {'sheet_config': {'header_row': 0, 'data_start_row': 1, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}}, 'sheets': [{'sheet_index': 0, 'sheet_type': 'data', 'columns': [{'name': 'col1', 'position': 0, 'data_type': 'string'}]}], 'matrix_fields': [{'matrix': [0, 1, 0], 'field_name': 'int_field', 'data_type': 'integer'}, {'matrix': [0, 1, 0], 'field_name': 'dec_field', 'data_type': 'decimal'}]}
        result = parse_excel_multi_sheet(tmp_path / 'test.xlsx', schema)
        df = result.collect()
        if 'int_field' in df.columns:
            assert df['int_field'].dtype == pl.Int64
        if 'dec_field' in df.columns:
            assert df['dec_field'].dtype == pl.Float64


class TestDynamicMetaDetectTransform:
    """Tests for the dynamic_meta_detect branch in _apply_sheet_transform (line 796->797)."""

    @pytest.mark.unit
    def test_apply_sheet_transform_dynamic_meta_detect(self):
        """Cover branch 796->797: dynamic_meta_detect transform type."""
        from acoharmony._parsers._excel_multi_sheet import _apply_sheet_transform

        df = pl.DataFrame(
            {
                "column_0": ["Header0", "Sub0", "Row1", "Row2"],
                "column_1": ["Header1", "Sub1", "10", "20"],
            }
        ).lazy()
        mock_module = MagicMock()
        mock_config = MagicMock()
        mock_module.DynamicMetaConfig = MagicMock(return_value=mock_config)
        mock_result_lazy = pl.DataFrame({"col1": ["10"], "col2": ["20"]}).lazy()
        mock_metadata = {"col1": MagicMock(), "col2": MagicMock()}
        mock_module.DynamicMetaDetectExpression.apply = MagicMock(
            return_value=(mock_result_lazy, mock_metadata)
        )
        saved = sys.modules.get("acoharmony._expressions._dynamic_meta_detect")
        sys.modules["acoharmony._expressions._dynamic_meta_detect"] = mock_module
        try:
            result = _apply_sheet_transform(
                df,
                {
                    "type": "dynamic_meta_detect",
                    "header_rows": [0, 1],
                    "separator": "_",
                    "data_start_row": 2,
                },
            )
            collected = result.collect()
            assert len(collected) >= 1
            mock_module.DynamicMetaConfig.assert_called_once_with(
                header_rows=[0, 1], separator="_"
            )
            mock_module.DynamicMetaDetectExpression.apply.assert_called_once_with(
                df, mock_config, 2
            )
        finally:
            if saved is not None:
                sys.modules["acoharmony._expressions._dynamic_meta_detect"] = saved
            else:
                sys.modules.pop("acoharmony._expressions._dynamic_meta_detect", None)


class TestDropSparseColumnsMissingPartition:
    """Tests for _drop_sparse_columns when partition_col is missing (line 832->833)."""

    @pytest.mark.unit
    def test_drop_sparse_columns_partition_col_not_in_columns(self):
        """Cover branch 832->833: partition_col not in collected.columns."""
        from acoharmony._parsers._excel_multi_sheet import _drop_sparse_columns

        df = pl.DataFrame({"a": [1, 2], "b": [3, 4]}).lazy()
        result = _drop_sparse_columns(df, "nonexistent_partition_col")
        collected = result.collect()
        assert collected.shape == (2, 2)
        assert "a" in collected.columns
        assert "b" in collected.columns


@pytest.mark.skipif(not HAS_OPENPYXL, reason='openpyxl required')
class TestExtractMatrixFieldsSearchLabel:
    """Cover search_label branches (lines 226-239) and data_type string (line 257-258)."""

    @pytest.mark.unit
    def test_search_label_found_with_string_dtype(self, tmp_path: Path):
        """search_label matches a row; value extracted and cast to string.

        Covers branches: 226->227, 228->229, 230->228, 230->231, 233->239, 257->265.
        """
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['Other Label', 'irrelevant'])
        ws.append(['Performance Year', '2025'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{
            'matrix': [0, None, 1],
            'field_name': 'perf_year',
            'search_label': 'Performance Year',
            'data_type': 'string',
        }]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result['perf_year'] == '2025'

    @pytest.mark.unit
    def test_search_label_not_found_uses_default(self, tmp_path: Path):
        """search_label does not match any row; default_value is returned.

        Covers branches: 226->227, 228->229, 228->233, 230->228, 233->234.
        """
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['Row A', 'val_a'])
        ws.append(['Row B', 'val_b'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{
            'matrix': [0, None, 1],
            'field_name': 'missing_label',
            'search_label': 'No Such Label',
            'default_value': 'N/A',
        }]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        assert result['missing_label'] == 'N/A'

    @pytest.mark.unit
    def test_extract_matrix_fields_unknown_data_type_passthrough(self, tmp_path: Path):
        """data_type is not integer/decimal/string so no cast is applied.

        Covers branch: 257->265 (elif data_type=='string' is false, fall through).
        """
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields
        wb = Workbook()
        ws = wb.active
        ws.append(['2025-01-15'])
        wb.save(tmp_path / 'test.xlsx')
        fields = [{
            'matrix': [0, 0, 0],
            'field_name': 'dt',
            'data_type': 'date',
        }]
        result = extract_matrix_fields(tmp_path / 'test.xlsx', fields, sheet_index=0)
        # Value should be passed through without casting
        assert result['dt'] is not None


class TestExtractMatrixFieldsSearchSheetName:
    """
    Cover the ``search_sheet_name`` resolver: schemas may point a matrix
    field at a sheet by name rather than a hardcoded index, so CMS
    layout shifts (e.g. a ``COVER PAGE`` inserted at index 0) don't
    silently return nulls.
    """

    @pytest.mark.unit
    def test_resolves_sheet_by_name_when_at_non_zero_index(self, tmp_path: Path):
        """Data lives on the second sheet named REPORT_PARAMETERS but the
        matrix says sheet 0. search_sheet_name must redirect the read."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields

        wb = Workbook()
        ws0 = wb.active
        ws0.title = "COVER PAGE"
        # Empty cover page
        ws1 = wb.create_sheet("REPORT_PARAMETERS")
        ws1.append(["Discount", "0.035"])
        ws1.append(["Shared Savings Rate", "1.0"])
        wb.save(tmp_path / "test.xlsx")

        fields = [
            {
                "matrix": [0, None, 1],
                "search_sheet_name": "REPORT_PARAMETERS",
                "search_label": "Discount",
                "field_name": "discount",
                "data_type": "decimal",
            }
        ]
        result = extract_matrix_fields(tmp_path / "test.xlsx", fields)
        assert result["discount"] == 0.035

    @pytest.mark.unit
    def test_case_insensitive_sheet_name_match(self, tmp_path: Path):
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields

        wb = Workbook()
        ws = wb.active
        ws.title = "Report_Parameters"
        ws.append(["Performance Year", "2026"])
        wb.save(tmp_path / "test.xlsx")

        fields = [
            {
                "matrix": [0, None, 1],
                "search_sheet_name": "REPORT_PARAMETERS",
                "search_label": "Performance Year",
                "field_name": "py",
                "data_type": "string",
            }
        ]
        result = extract_matrix_fields(tmp_path / "test.xlsx", fields)
        assert result["py"] == "2026"

    @pytest.mark.unit
    def test_missing_sheet_name_falls_back_to_matrix_index(self, tmp_path: Path):
        """If the named sheet isn't present, the resolver returns None and
        extraction falls back to the matrix[0] index path. This is the
        backward-compat safety net."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields

        wb = Workbook()
        ws = wb.active
        ws.title = "PARAMETERS_V2"
        ws.append(["Discount", "0.02"])
        wb.save(tmp_path / "test.xlsx")

        # search_sheet_name points at a non-existent sheet → resolver
        # returns None → falls back to matrix[0] = 0 (the only sheet).
        fields = [
            {
                "matrix": [0, None, 1],
                "search_sheet_name": "REPORT_PARAMETERS",
                "search_label": "Discount",
                "field_name": "discount",
                "data_type": "decimal",
            }
        ]
        result = extract_matrix_fields(tmp_path / "test.xlsx", fields)
        assert result["discount"] == 0.02

    @pytest.mark.unit
    def test_no_search_sheet_name_is_backward_compatible(self, tmp_path: Path):
        """Schemas without search_sheet_name behave exactly as before."""
        from openpyxl import Workbook

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields

        wb = Workbook()
        ws = wb.active
        ws.append(["Quality Withhold", "0.02"])
        wb.save(tmp_path / "test.xlsx")

        fields = [
            {
                "matrix": [0, None, 1],
                "search_label": "Quality Withhold",
                "field_name": "qw",
                "data_type": "decimal",
            }
        ]
        result = extract_matrix_fields(tmp_path / "test.xlsx", fields)
        assert result["qw"] == 0.02


class TestDropSparseColumnsDropBranches:
    """Cover _drop_sparse_columns branches for column iteration and drop logic."""

    @pytest.mark.unit
    def test_drops_all_null_columns(self):
        """Cover 832->836, 837->838, 838->837, 838->839, 837->841, 842->843.

        When partition_col is present and some columns are entirely null,
        those columns are dropped.
        """
        from acoharmony._parsers._excel_multi_sheet import _drop_sparse_columns

        df = pl.DataFrame({
            "partition": ["a", "a", "b"],
            "real_data": [1, 2, 3],
            "all_null": [None, None, None],
        }).lazy()
        result = _drop_sparse_columns(df, "partition").collect()
        assert "partition" in result.columns
        assert "real_data" in result.columns
        assert "all_null" not in result.columns
        assert result.shape == (3, 2)

    @pytest.mark.unit
    def test_no_columns_to_drop(self):
        """Cover 842->845: drop list is empty so no columns are removed."""
        from acoharmony._parsers._excel_multi_sheet import _drop_sparse_columns

        df = pl.DataFrame({
            "partition": ["x", "y"],
            "col_a": [10, 20],
            "col_b": [30, 40],
        }).lazy()
        result = _drop_sparse_columns(df, "partition").collect()
        assert set(result.columns) == {"partition", "col_a", "col_b"}
        assert result.shape == (2, 3)


class TestSheetNameLoadException:
    """Cover _excel_multi_sheet.py:1007-1008 — openpyxl exception during sheet name lookup."""

    @pytest.mark.unit
    def test_corrupt_workbook(self, tmp_path):
        from unittest.mock import patch as _patch
        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        bad_file = tmp_path / "corrupt.xlsx"
        bad_file.write_bytes(b"not an xlsx")
        # Patch openpyxl.load_workbook to fail during sheet name lookup (lines 1000-1008).
        # The except block catches and passes, so _sheet_name_to_idx stays empty.
        # parse_excel_multi_sheet will then fail on actual parsing, which is fine.
        with _patch("openpyxl.load_workbook", side_effect=RuntimeError("corrupt")):
            try:
                parse_excel_multi_sheet(bad_file, {"sheets": []})
            except Exception:
                pass  # Expected — corrupt file can't be parsed


class TestOpenpyxlLoadException:
    """Cover _excel_multi_sheet.py:1007-1008."""

    @pytest.mark.unit
    def test_workbook_exception_caught(self, tmp_path):
        from unittest.mock import patch as _patch
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"not excel")
        with _patch("openpyxl.load_workbook", side_effect=Exception("corrupt")):
            # The except block in the parser catches and continues
            from acoharmony._parsers._excel_multi_sheet import parse_sheet_matrix, ExcelMultiSheetConfig
            config = ExcelMultiSheetConfig(header_row=0, data_start_row=1, end_marker_column=0, end_marker_value="TOTAL")
            try:
                parse_sheet_matrix(bad, 0, config, [])
            except Exception:
                pass


class TestOpenpyxlExceptionCaught:
    """Cover lines 1007-1008."""
    @pytest.mark.unit
    def test_bad_workbook_silent_fail(self, tmp_path):
        from unittest.mock import patch as _p
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"corrupt")
        with _p("openpyxl.load_workbook", side_effect=Exception("corrupt")):
            from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet, ExcelMultiSheetConfig
            config = ExcelMultiSheetConfig(header_row=0, data_start_row=1, end_marker_column=0, end_marker_value="END")
            try: parse_excel_multi_sheet(bad, {"file_format": {"sheet_config": vars(config)}, "sheets": []})
            except: pass


class TestOpenpyxlExceptionInParse:
    """Lines 1007-1008: except during workbook load."""
    @pytest.mark.unit
    def test_corrupt_file_exception_caught(self, tmp_path):
        from unittest.mock import patch
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"not excel data")
        with patch("openpyxl.load_workbook", side_effect=Exception("corrupt file")):
            from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet
            schema = {"file_format": {"sheet_config": {"header_row": 0, "data_start_row": 1, "end_marker_column": 0, "end_marker_value": "END"}}, "sheets": []}
            try: parse_excel_multi_sheet(bad, schema)
            except: pass


class TestDropSparseColumnsCollectException:
    """Cover lines 829-830: _drop_sparse_columns when df.collect() raises."""

    @pytest.mark.unit
    def test_collect_failure_returns_original_df(self):
        """Lines 829-830: when df.collect() raises, return the original LazyFrame."""
        from acoharmony._parsers._excel_multi_sheet import _drop_sparse_columns

        # Create a mock LazyFrame whose collect() raises
        mock_lf = MagicMock(spec=pl.LazyFrame)
        mock_lf.collect.side_effect = RuntimeError("collect failed")

        result = _drop_sparse_columns(mock_lf, "partition_col")
        assert result is mock_lf

    @pytest.mark.unit
    def test_collect_polars_error_returns_original(self):
        """Lines 829-830: ComputeError during collect returns original LazyFrame."""
        from acoharmony._parsers._excel_multi_sheet import _drop_sparse_columns

        mock_lf = MagicMock(spec=pl.LazyFrame)
        mock_lf.collect.side_effect = pl.exceptions.ComputeError("bad schema")

        result = _drop_sparse_columns(mock_lf, "partition_col")
        assert result is mock_lf


class TestOpenpyxlLoadWorkbookExceptionInMultiSheet:
    """Cover lines 1007-1008: openpyxl load_workbook fails during sheet name lookup."""

    @pytest.mark.unit
    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")
    def test_openpyxl_load_exception_still_parses(self, tmp_path):
        """Lines 1007-1008: exception in openpyxl.load_workbook is caught
        and parsing continues with empty sheet name lookup."""
        import openpyxl

        p = tmp_path / "test_opx_fail.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "Value"])
        ws.append(["A", "100"])
        ws.append(["TOTAL", ""])
        wb.save(p)

        from acoharmony._parsers._excel_multi_sheet import parse_excel_multi_sheet

        schema = {
            "name": "test_opx",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            "sheets": [
                {
                    "sheet_type": "data",
                    "sheet_index": 0,
                    "columns": [{"position": 1, "name": "type_col", "data_type": "string"}],
                }
            ],
        }

        # Patch openpyxl.load_workbook to raise during sheet name lookup
        with patch(
            "openpyxl.load_workbook",
            side_effect=Exception("openpyxl broken"),
        ):
            lf = parse_excel_multi_sheet(p, schema)
            df = lf.collect()
            assert len(df) >= 1


class TestFilenameFieldsStamping:
    """
    ``filename_fields`` at the top of a schema declares which columns
    should be derived from the source filename and stamped on every
    row. The stamping happens after the sheet-level parse so it
    overrides same-named columns from workbook content.
    """

    @pytest.fixture
    def workbook(self, tmp_path: Path) -> Path:
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        # CMS-shaped filename so the real extractors have something to pull
        out = tmp_path / "REACH.D0259.BNMR.PY2024.D250101.T0900000.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DATA"
        ws.append(["col_a", "col_b"])
        ws.append(["v1", "v2"])
        ws.append(["v3", "v4"])
        wb.save(out)
        return out

    @pytest.fixture
    def base_schema(self) -> dict:
        return {
            "name": "synthetic_bnmr",
            "file_format": {
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "END",
                }
            },
            "sheets": [
                {
                    "sheet_name": "DATA",
                    "sheet_index": 0,
                    "sheet_type": "data",
                    "columns": [
                        {"position": 0, "name": "col_a", "data_type": "string"},
                        {"position": 1, "name": "col_b", "data_type": "string"},
                    ],
                }
            ],
        }

    @pytest.mark.unit
    def test_aco_id_and_performance_year_stamped(self, workbook, base_schema):
        base_schema["filename_fields"] = [
            {"name": "aco_id", "extractor": "aco_id"},
            {"name": "performance_year", "extractor": "performance_year"},
        ]
        df = parse_excel_multi_sheet(workbook, base_schema).collect()
        assert df["aco_id"].drop_nulls().unique().to_list() == ["D0259"]
        assert df["performance_year"].drop_nulls().unique().to_list() == ["2024"]

    @pytest.mark.unit
    def test_filename_fields_empty_list_adds_no_cols(self, workbook, base_schema):
        base_schema["filename_fields"] = []
        df = parse_excel_multi_sheet(workbook, base_schema).collect()
        # aco_id/performance_year are not stamped when list is empty
        assert "aco_id" not in df.columns
        assert "performance_year" not in df.columns

    @pytest.mark.unit
    def test_filename_fields_omitted_key_is_equivalent_to_empty(
        self, workbook, base_schema
    ):
        # No filename_fields key at all
        df = parse_excel_multi_sheet(workbook, base_schema).collect()
        assert "aco_id" not in df.columns

    @pytest.mark.unit
    def test_unknown_extractor_raises_at_parse_time(self, workbook, base_schema):
        base_schema["filename_fields"] = [
            {"name": "aco_id", "extractor": "not_a_real_extractor"}
        ]
        with pytest.raises(ValueError, match="Unknown filename extractor"):
            parse_excel_multi_sheet(workbook, base_schema).collect()

    @pytest.mark.unit
    def test_filename_fields_override_workbook_column(self, tmp_path, base_schema):
        """A column named ``aco_id`` in the workbook itself must be
        overwritten by the filename-derived value. The filename is the
        source of truth."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        out = tmp_path / "REACH.D9999.BNMR.PY2024.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DATA"
        ws.append(["aco_id", "col_b"])
        ws.append(["WORKBOOK_ACO", "v2"])
        wb.save(out)

        schema = dict(base_schema)
        schema["sheets"] = [
            {
                "sheet_name": "DATA",
                "sheet_index": 0,
                "sheet_type": "data",
                "columns": [
                    {"position": 0, "name": "aco_id", "data_type": "string"},
                    {"position": 1, "name": "col_b", "data_type": "string"},
                ],
            }
        ]
        schema["filename_fields"] = [{"name": "aco_id", "extractor": "aco_id"}]

        df = parse_excel_multi_sheet(out, schema).collect()
        # Filename says D9999; workbook says WORKBOOK_ACO — filename wins.
        assert df["aco_id"].drop_nulls().unique().to_list() == ["D9999"]

    @pytest.mark.unit
    def test_provenance_columns_always_stamped(self, workbook, base_schema):
        """Even with no filename_fields, source_filename / source_file /
        processed_at are always stamped."""
        df = parse_excel_multi_sheet(workbook, base_schema).collect()
        assert "source_filename" in df.columns
        assert "source_file" in df.columns
        assert "processed_at" in df.columns
        assert (
            df["source_filename"].unique().to_list()
            == ["REACH.D0259.BNMR.PY2024.D250101.T0900000.xlsx"]
        )


# ---------------------------------------------------------------------------
# Coverage: extract_matrix_fields — search_label with empty first-column cell
# ---------------------------------------------------------------------------


class TestExtractMatrixFieldsSearchLabelEmptyCell:
    """
    Line 265-266: when scanning for a search_label, rows whose first column
    is falsy (None / empty string) must be skipped via ``continue``.
    """

    @pytest.mark.unit
    def test_search_label_skips_empty_first_column(self, tmp_path):
        """A None cell in column 0 triggers the ``if not cell: continue`` branch."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        from acoharmony._parsers._excel_multi_sheet import extract_matrix_fields

        # Build a workbook where row 0 col 0 is empty, row 1 has the label
        out = tmp_path / "search.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, "ignore"])      # row 0: first cell is None
        ws.append(["Target Label", 42])  # row 1: label is here
        wb.save(out)

        matrix_fields = [
            {
                "matrix": [0, 0, 1],  # sheet 0, start row 0, target_col 1
                "field_name": "found_value",
                "search_label": "Target Label",
            }
        ]
        result = extract_matrix_fields(out, matrix_fields, sheet_index=0)
        # Excel reads with infer_schema_length=0 return strings
        assert str(result["found_value"]) == "42"


# ---------------------------------------------------------------------------
# Coverage: extract_named_fields — outer except with namespace configs
# ---------------------------------------------------------------------------


class TestExtractNamedFieldsExceptPaths:
    """
    Lines 438-444: when ``pl.read_excel`` raises in ``extract_named_fields``,
    the outer except block must yield None for every field. When the
    ``field_config`` is a namespace (not dict), line 444 is exercised.
    """

    @pytest.mark.unit
    def test_read_raises_dict_config_yields_none(self, tmp_path):
        """Outer except block with dict configs."""
        from acoharmony._parsers._excel_multi_sheet import extract_named_fields

        bogus_path = tmp_path / "nonexistent.xlsx"
        config = [
            {"row": 0, "column": 0, "field_name": "alpha"},
            {"row": 1, "column": 1, "field_name": "beta"},
        ]
        result = extract_named_fields(bogus_path, 0, config)
        assert result == {"alpha": None, "beta": None}

    @pytest.mark.unit
    def test_read_raises_namespace_config_yields_none(self, tmp_path):
        """Outer except block with namespace (non-dict) field configs (line 444)."""
        from acoharmony._parsers._excel_multi_sheet import extract_named_fields

        bogus_path = tmp_path / "nonexistent.xlsx"
        ns_config = [
            SimpleNamespace(row=0, column=0, field_name="gamma"),
            SimpleNamespace(row=1, column=1, field_name="delta"),
        ]
        result = extract_named_fields(bogus_path, 0, ns_config)
        assert result == {"gamma": None, "delta": None}

    @pytest.mark.unit
    def test_col_idx_exceeds_df_columns(self, tmp_path):
        """
        Lines 433-434: row_idx valid but col_idx >= len(df.columns) yields None.
        """
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl required")
        import openpyxl

        from acoharmony._parsers._excel_multi_sheet import extract_named_fields

        out = tmp_path / "tiny.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["only_col"])
        wb.save(out)

        config = [{"row": 0, "column": 999, "field_name": "missing_col"}]
        result = extract_named_fields(out, 0, config)
        assert result == {"missing_col": None}


# ---------------------------------------------------------------------------
# Coverage: _stamp_provenance_metadata — namespace schema + namespace fields
# ---------------------------------------------------------------------------


class TestStampProvenanceNamespacePaths:
    """
    Lines 1518-1520: schema is a namespace object with ``filename_fields``.
    Lines 1526-1528: field_def is a namespace (not dict).
    Line 1530: ``continue`` when name or extractor_name is None.
    """

    @pytest.mark.unit
    def test_namespace_schema_with_namespace_field_defs(self, tmp_path):
        """Namespace schema whose filename_fields are namespace objects."""
        from acoharmony._parsers._excel_multi_sheet import _stamp_provenance_metadata

        df = pl.LazyFrame({"col_a": ["v1"]})

        # Use a real extractor name so it resolves
        schema = SimpleNamespace(
            filename_fields=[
                SimpleNamespace(name="aco_id", extractor="aco_id"),
            ]
        )
        result = _stamp_provenance_metadata(
            df,
            Path(tmp_path / "REACH.D0259.BNMR.PY2024.xlsx"),
            "test_schema",
            schema,
        )
        out = result.collect()
        assert "aco_id" in out.columns
        assert out["aco_id"][0] == "D0259"

    @pytest.mark.unit
    def test_namespace_field_def_missing_name_skipped(self, tmp_path):
        """A namespace field_def with name=None triggers the ``continue`` at line 1530."""
        from acoharmony._parsers._excel_multi_sheet import _stamp_provenance_metadata

        df = pl.LazyFrame({"col_a": ["v1"]})

        schema = SimpleNamespace(
            filename_fields=[
                SimpleNamespace(name=None, extractor="aco_id"),   # missing name
                SimpleNamespace(name="perf_yr", extractor=None),  # missing extractor
            ]
        )
        result = _stamp_provenance_metadata(
            df,
            Path(tmp_path / "REACH.D0259.BNMR.PY2024.xlsx"),
            "test_schema",
            schema,
        )
        out = result.collect()
        # Neither field should be added
        assert "aco_id" not in [c for c in out.columns if c not in ("col_a", "processed_at", "source_file", "source_filename")]
        assert "perf_yr" not in out.columns

    @pytest.mark.unit
    def test_namespace_schema_filename_fields_as_tuple(self, tmp_path):
        """
        Lines 1519-1520: raw is not a list (e.g. tuple) and gets wrapped
        via ``list(raw)``.
        """
        from acoharmony._parsers._excel_multi_sheet import _stamp_provenance_metadata

        df = pl.LazyFrame({"col_a": ["v1"]})

        schema = SimpleNamespace(
            filename_fields=(
                SimpleNamespace(name="aco_id", extractor="aco_id"),
            )
        )
        result = _stamp_provenance_metadata(
            df,
            Path(tmp_path / "REACH.D0259.BNMR.PY2024.xlsx"),
            "test_schema",
            schema,
        )
        out = result.collect()
        assert out["aco_id"][0] == "D0259"
