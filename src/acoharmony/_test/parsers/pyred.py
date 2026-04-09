from __future__ import annotations

# Magic auto-import: brings in ALL exports from module under test
from acoharmony._test._import_magic import auto_import


@auto_import
class _:
    pass  # noqa: E701

from unittest.mock import patch
from pathlib import Path
from typing import TYPE_CHECKING
from types import SimpleNamespace

import polars as pl
import pytest
import acoharmony

from acoharmony._parsers._pyred import parse_pyred
from acoharmony.tables import TableManager

from .conftest import HAS_OPENPYXL

'Tests for acoharmony._parsers._pyred module.'

class TestModuleStructure:
    """Basic module structure tests."""

    @pytest.mark.unit
    def test_module_imports(self):
        """Module can be imported."""
        assert acoharmony._parsers._pyred is not None
'Unit tests for _pyred module.'
if TYPE_CHECKING:
    pass


@pytest.mark.unit
def test_extract_pyred_metadata_basic() -> None:
    """extract_pyred_metadata basic functionality."""
    from acoharmony._parsers._pyred import extract_pyred_metadata
    # Calling with a nonexistent file returns None values gracefully
    result = extract_pyred_metadata(Path('/nonexistent_file.xlsx'), 0)
    assert isinstance(result, dict)
    assert 'performance_year' in result
    assert 'report_period' in result
    assert result['performance_year'] is None
    assert result['report_period'] is None

@pytest.mark.unit
def test_parse_pyred_basic() -> None:
    """parse_pyred basic functionality -- bad schema raises ValueError."""
    from acoharmony._parsers._pyred import parse_pyred
    with pytest.raises(ValueError, match='file_format'):
        parse_pyred(Path('/tmp/nonexistent.xlsx'), {'name': 'x', 'sheets': []})

@pytest.mark.unit
def test_parse_pyred_sheets_list_none():
    """Line 139: sheets_list defaults to [] when None."""
    schema = type('Schema', (), {'file_format': type('FF', (), {'sheet_config': {'Sheet1': {}}})(), 'sheets': None})()
    try:
        parse_pyred(Path('/tmp/nonexistent.xlsx'), schema)
    except (FileNotFoundError, ValueError, AttributeError):
        pass

class TestParsePyred:
    """Tests for _pyred.extract_pyred_metadata and parse_pyred."""

    @pytest.fixture
    def pyred_xlsx(self, tmp_path: Path) -> Path:
        if not HAS_OPENPYXL:
            pytest.skip('openpyxl required')
        import openpyxl
        p = tmp_path / 'pyred.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Physician'
        ws.append(['Performance Year 2025', None, None, None, None])
        ws.append([None, None, None, None, 'January 2025 experience'])
        ws.append([None, None, None, None, None])
        ws.append(['Provider Type', 'NPI', 'Name', 'Amount', 'Status'])
        ws.append(['Physician', '1234567890', 'Dr. Smith', '100.00', 'Active'])
        ws.append(['Physician', '0987654321', 'Dr. Jones', '200.00', 'Active'])
        ws.append(['TOTAL', '', '', '300.00', ''])
        wb.save(p)
        return p

    @pytest.mark.unit
    def test_extract_pyred_metadata(self, pyred_xlsx: Path):
        from acoharmony._parsers._pyred import extract_pyred_metadata
        result = extract_pyred_metadata(pyred_xlsx, 0)
        assert result['performance_year'] == '2025'
        assert result['report_period'] is not None
        assert '2025' in result['report_period']

    @pytest.mark.unit
    def test_extract_pyred_metadata_not_found(self, tmp_path: Path):
        """When metadata cells don't match patterns, return None."""
        df = pl.DataFrame({'A': ['nothing here'], 'B': ['also nothing']})
        p = tmp_path / 'empty_meta.xlsx'
        df.write_excel(p)
        from acoharmony._parsers._pyred import extract_pyred_metadata
        result = extract_pyred_metadata(p, 0)
        assert result['performance_year'] is None
        assert result['report_period'] is None

    @pytest.mark.unit
    def test_extract_pyred_metadata_exception(self, tmp_path: Path):
        """When file can't be read, return None values."""
        from acoharmony._parsers._pyred import extract_pyred_metadata
        result = extract_pyred_metadata(Path('/nonexistent.xlsx'), 0)
        assert result['performance_year'] is None
        assert result['report_period'] is None

    @pytest.mark.unit
    def test_parse_pyred_basic(self, pyred_xlsx: Path):
        from acoharmony._parsers._pyred import parse_pyred
        schema = {'name': 'pyred', 'file_format': {'sheet_config': {'header_row': 3, 'data_start_row': 4, 'end_marker_column': 0, 'end_marker_value': 'TOTAL', 'column_mapping_strategy': 'position'}}, 'sheets': [{'sheet_type': 'physician', 'sheet_index': 0, 'columns': [{'position': 0, 'name': 'provider_type', 'data_type': 'string'}, {'position': 1, 'name': 'npi', 'data_type': 'string'}, {'position': 2, 'name': 'name', 'data_type': 'string'}, {'position': 3, 'name': 'amount', 'data_type': 'string'}]}]}
        lf = parse_pyred(pyred_xlsx, schema)
        df = lf.collect()
        assert 'sheet_type' in df.columns
        assert 'performance_year' in df.columns
        assert 'report_period' in df.columns
        assert len(df) >= 1

    @pytest.mark.unit
    def test_parse_pyred_sheet_filter(self, pyred_xlsx: Path):
        from acoharmony._parsers._pyred import parse_pyred
        schema = {'name': 'pyred', 'file_format': {'sheet_config': {'header_row': 3, 'data_start_row': 4, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}}, 'sheets': [{'sheet_type': 'physician', 'sheet_index': 0, 'columns': [{'position': 0, 'name': 'provider_type', 'data_type': 'string'}]}]}
        with pytest.raises(ValueError, match='No sheets found'):
            parse_pyred(pyred_xlsx, schema, sheet_types=['inpatient'])

    @pytest.mark.unit
    def test_parse_pyred_bad_schema_no_file_format(self, tmp_path: Path):
        p = tmp_path / 'x.xlsx'
        pl.DataFrame({'a': [1]}).write_excel(p)
        from acoharmony._parsers._pyred import parse_pyred
        with pytest.raises(ValueError, match='file_format'):
            parse_pyred(p, {'name': 'x', 'sheets': []})

    @pytest.mark.unit
    def test_parse_pyred_bad_schema_no_sheet_config(self, tmp_path: Path):
        p = tmp_path / 'x.xlsx'
        pl.DataFrame({'a': [1]}).write_excel(p)
        from acoharmony._parsers._pyred import parse_pyred
        with pytest.raises(ValueError, match='sheet_config'):
            parse_pyred(p, {'file_format': {}, 'sheets': [{'sheet_type': 'a'}]})

    @pytest.mark.unit
    def test_parse_pyred_bad_schema_no_sheets(self, tmp_path: Path):
        p = tmp_path / 'x.xlsx'
        pl.DataFrame({'a': [1]}).write_excel(p)
        from acoharmony._parsers._pyred import parse_pyred
        with pytest.raises(ValueError, match='sheets'):
            parse_pyred(p, {'file_format': {'sheet_config': {'header_row': 0, 'data_start_row': 1, 'end_marker_column': 0, 'end_marker_value': 'END'}}, 'sheets': []})

    @pytest.mark.unit
    def test_parse_pyred_with_limit(self, pyred_xlsx: Path):
        from acoharmony._parsers._pyred import parse_pyred
        schema = {'name': 'pyred', 'file_format': {'sheet_config': {'header_row': 3, 'data_start_row': 4, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}}, 'sheets': [{'sheet_type': 'physician', 'sheet_index': 0, 'columns': [{'position': 0, 'name': 'provider_type', 'data_type': 'string'}]}]}
        lf = parse_pyred(pyred_xlsx, schema, limit=1)
        df = lf.collect()
        assert len(df) <= 1

    @pytest.mark.unit
    def test_parse_pyred_namespace_schema(self, pyred_xlsx: Path):
        """Test with SimpleNamespace schema (non-dict)."""
        from acoharmony._parsers._pyred import parse_pyred
        schema = SimpleNamespace(name='pyred', file_format=SimpleNamespace(sheet_config={'header_row': 3, 'data_start_row': 4, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}), sheets=[SimpleNamespace(sheet_type='physician', sheet_index=0, columns=[SimpleNamespace(position=0, name='provider_type', data_type='string')])])
        lf = parse_pyred(pyred_xlsx, schema)
        df = lf.collect()
        assert len(df) >= 1

class TestPyredAdditional:
    """Additional tests to fill coverage gaps in _pyred.py."""

    @pytest.mark.unit
    def test_parse_pyred_namespace_schema_with_non_list_sheets(self, tmp_path: Path):
        """Cover lines 139-143: namespace schema where sheets is a tuple."""
        if not HAS_OPENPYXL:
            pytest.skip('openpyxl required')
        import openpyxl
        p = tmp_path / 'pyred_ns.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Performance Year 2025', None, None, None, None])
        ws.append([None, None, None, None, 'January 2025 experience'])
        ws.append([None])
        ws.append(['Provider Type', 'NPI', 'Name'])
        ws.append(['Physician', '123', 'Dr. X'])
        ws.append(['TOTAL', '', ''])
        wb.save(p)
        from acoharmony._parsers._pyred import parse_pyred
        schema = SimpleNamespace(name='pyred', file_format=SimpleNamespace(sheet_config={'header_row': 3, 'data_start_row': 4, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}), sheets=(SimpleNamespace(sheet_type='physician', sheet_index=0, columns=[SimpleNamespace(position=2, name='name', data_type='string')]),))
        lf = parse_pyred(p, schema)
        df = lf.collect()
        assert 'sheet_type' in df.columns

    @pytest.mark.unit
    def test_parse_pyred_namespace_sheet_config_non_dict(self, tmp_path: Path):
        """Cover line 154: sheet_config_dict from namespace."""
        if not HAS_OPENPYXL:
            pytest.skip('openpyxl required')
        import openpyxl
        p = tmp_path / 'pyred_sc.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Performance Year 2025'])
        ws.append([None])
        ws.append([None])
        ws.append(['Provider Type', 'NPI', 'Name'])
        ws.append(['Physician', '123', 'Dr. X'])
        ws.append(['TOTAL', '', ''])
        wb.save(p)
        from acoharmony._parsers._pyred import parse_pyred
        schema = {'name': 'pyred', 'file_format': {'sheet_config': SimpleNamespace(header_row=3, data_start_row=4, end_marker_column=0, end_marker_value='TOTAL', column_mapping_strategy='position', header_search_text=None)}, 'sheets': [{'sheet_type': 'physician', 'sheet_index': 0, 'columns': [{'position': 2, 'name': 'name', 'data_type': 'string'}]}]}
        lf = parse_pyred(p, schema)
        df = lf.collect()
        assert 'sheet_type' in df.columns

    @pytest.mark.unit
    def test_parse_pyred_header_metadata_deduplication(self, tmp_path: Path):
        """Cover lines 189-194: header metadata added only if not existing."""
        if not HAS_OPENPYXL:
            pytest.skip('openpyxl required')
        import openpyxl
        p = tmp_path / 'pyred_hm.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Performance Year 2025'])
        ws.append([None])
        ws.append([None])
        ws.append(['Provider Type', 'Amount'])
        ws.append(['Physician', '200'])
        ws.append(['TOTAL', ''])
        wb.save(p)
        from acoharmony._parsers._pyred import parse_pyred
        mock_lf = pl.DataFrame({'provider_type': ['Physician']}).lazy()
        mock_meta = {'provider_type': {'claim_year': '2025'}}
        with patch('acoharmony._parsers._pyred.parse_sheet_matrix', return_value=(mock_lf, mock_meta)):
            schema = {'name': 'pyred', 'file_format': {'sheet_config': {'header_row': 3, 'data_start_row': 4, 'end_marker_column': 0, 'end_marker_value': 'TOTAL'}}, 'sheets': [{'sheet_type': 'physician', 'sheet_index': 0, 'columns': [{'position': 1, 'name': 'provider_type', 'data_type': 'string'}]}]}
            lf = parse_pyred(p, schema)
            df = lf.collect()
            assert 'claim_year' in df.columns

@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")
class TestPyredCoverageGaps:
    """Cover _pyred.py missed lines 139, 143."""

    @pytest.mark.unit
    def test_pyred_schema_object_non_list_sheets(self, tmp_path: Path):
        """Cover line 139-141: sheets is not a list → converted, then fails on empty."""
        from acoharmony._parsers._pyred import parse_pyred

        schema = SimpleNamespace(
            file_format={
                "sheet_config": {
                    "header_row": 0,
                    "data_start_row": 1,
                    "end_marker_column": 0,
                    "end_marker_value": "TOTAL",
                }
            },
            sheets=(),
        )
        with pytest.raises(ValueError, match="sheets"):
            parse_pyred(tmp_path / "test.xlsx", schema)

    @pytest.mark.unit
    def test_pyred_no_file_format_raises(self, tmp_path: Path):
        """Cover line 143: no file_format → ValueError."""
        from acoharmony._parsers._pyred import parse_pyred

        schema = SimpleNamespace(name="test")
        with pytest.raises(ValueError, match="file_format"):
            parse_pyred(tmp_path / "test.xlsx", schema)


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")
class TestPyredPerformanceYearLoopBackEdge:
    """Cover branch 56→52: 'Performance Year' present but no 4-digit year."""

    @pytest.mark.unit
    def test_extract_metadata_performance_year_no_digits(self, tmp_path: Path):
        """When a row contains 'Performance Year' but no 4-digit number,
        the regex match is None so the loop continues to the next row
        (branch 56→52)."""
        import openpyxl
        from acoharmony._parsers._pyred import extract_pyred_metadata

        p = tmp_path / "pyred_no_year.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Row 0: contains "Performance Year" but NO 4-digit number
        ws.append(["Performance Year TBD"])
        # Row 1: contains "Performance Year" WITH a 4-digit number (loop finds it)
        ws.append(["Performance Year 2025"])
        ws.append([None])
        ws.append(["Header"])
        wb.save(p)

        result = extract_pyred_metadata(p, 0)
        # First row triggers `if cell_value and "Performance Year" in str(cell_value)`
        # but re.search(r"\d{4}", ...) returns None → loop continues (56→52).
        # Second row matches both conditions → performance_year = "2025".
        assert result["performance_year"] == "2025"


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")
class TestPyredReportPeriodLoopBackEdges:
    """Cover branch 72->67 and 74->67: report-period search_locations loop back-edges."""

    @pytest.mark.unit
    def test_report_period_first_location_no_experience_keyword(self, tmp_path: Path):
        """Line 72->67: first search location has a non-None cell value that does NOT
        contain 'experience', so the loop continues to the second location."""
        import openpyxl
        from acoharmony._parsers._pyred import extract_pyred_metadata

        p = tmp_path / "pyred_72_67.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Row 0: performance year
        ws.append(["Performance Year 2025", "x", "x", "x", "x", "x", "x", "x"])
        # Row 1 (search_locations[0] = (1,4)): col 4 has text but NOT 'experience'
        ws.append(["r1", "x", "x", "x", "Some Other Text", "x", "x", "x"])
        # Row 2: filler (non-empty to prevent row collapsing)
        ws.append(["r2", "x", "x", "x", "x", "x", "x", "x"])
        # Row 3 (search_locations[1] = (3,7)): col 7 has 'experience' with valid date
        ws.append(["r3", "x", "x", "x", "x", "x", "x", "January 2025 experience"])
        wb.save(p)

        result = extract_pyred_metadata(p, 0)
        # First location (1,4) has "Some Other Text" -> no 'experience' -> loop back to 67
        # Second location (3,7) has "January 2025 experience" -> match -> report_period set
        assert result["report_period"] == "January 2025"

    @pytest.mark.unit
    def test_report_period_experience_but_no_date_pattern(self, tmp_path: Path):
        """Line 74->67: first search location has 'experience' in the cell but the
        regex for 'Month YYYY' does NOT match, so the loop continues."""
        import openpyxl
        from acoharmony._parsers._pyred import extract_pyred_metadata

        p = tmp_path / "pyred_74_67.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Row 0: performance year
        ws.append(["Performance Year 2025", "x", "x", "x", "x", "x", "x", "x"])
        # Row 1 (search_locations[0] = (1,4)): col 4 has 'experience' but no Month YYYY
        ws.append(["r1", "x", "x", "x", "experience data only", "x", "x", "x"])
        # Row 2: filler (non-empty to prevent row collapsing)
        ws.append(["r2", "x", "x", "x", "x", "x", "x", "x"])
        # Row 3 (search_locations[1] = (3,7)): col 7 has valid 'experience' with date
        ws.append(["r3", "x", "x", "x", "x", "x", "x", "March 2025 experience"])
        wb.save(p)

        result = extract_pyred_metadata(p, 0)
        # First location: has 'experience' but "experience data only" has no "Month YYYY"
        # -> regex returns None -> loop back to 67
        # Second location: "March 2025 experience" matches -> report_period = "March 2025"
        assert result["report_period"] == "March 2025"


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")
class TestPyredMetadataMultipleFields:
    """Cover branch 190->189: inner loop iterates over multiple fields in metadata_dict."""

    @pytest.mark.unit
    def test_col_header_metadata_multiple_fields_per_column(self, tmp_path: Path):
        """Line 190->189: col_header_metadata has a column entry with multiple
        field_name/field_value pairs, so the inner for-loop iterates back to 189."""
        import openpyxl
        from acoharmony._parsers._pyred import parse_pyred

        p = tmp_path / "pyred_190_189.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Performance Year 2025"])
        ws.append([None])
        ws.append([None])
        ws.append(["Col A", "Col B"])
        ws.append(["val1", "val2"])
        ws.append(["TOTAL", ""])
        wb.save(p)

        mock_lf = pl.DataFrame({"col_a": ["val1"]}).lazy()
        # The inner loop must skip a field_name that already exists in df columns
        # (line 190 False branch -> back to 189) then process the next field.
        # "col_a" already exists in the DataFrame, so it's skipped; "new_field" is added.
        mock_meta = {
            "col_a": {"col_a": "duplicate_ignored", "new_field": "new_value"},
        }
        with patch(
            "acoharmony._parsers._pyred.parse_sheet_matrix",
            return_value=(mock_lf, mock_meta),
        ):
            schema = {
                "name": "pyred",
                "file_format": {
                    "sheet_config": {
                        "header_row": 3,
                        "data_start_row": 4,
                        "end_marker_column": 0,
                        "end_marker_value": "TOTAL",
                    }
                },
                "sheets": [
                    {
                        "sheet_type": "physician",
                        "sheet_index": 0,
                        "columns": [{"position": 0, "name": "col_a", "data_type": "string"}],
                    }
                ],
            }
            lf = parse_pyred(p, schema)
            df = lf.collect()
            # "col_a" already existed -> skipped (190->189 back-edge)
            # "new_field" did not exist -> added as a new column
            assert "new_field" in df.columns
            assert df["new_field"][0] == "new_value"
            # col_a should still be the original value, not overwritten
            assert df["col_a"][0] == "val1"
